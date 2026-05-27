#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Wealth Engine Streamlit UI FINAL - Anaconda/local friendly."""
from __future__ import annotations

import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
import base64
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    load_workbook = None

st.set_page_config(page_title="Wealth Engine", page_icon="💰", layout="wide", initial_sidebar_state="expanded")

APP_TITLE = "Wealth Engine"
BASE_DIR = Path("/content") if Path("/content").exists() else Path.cwd()
CORE_SCRIPT_CANDIDATES = [
    BASE_DIR / "wealth_engine.py",
    BASE_DIR / "Wealth Engine.py",
    BASE_DIR / "Wealth_Engine.py",
    BASE_DIR / "WealthEngine.py",
    BASE_DIR / "script.py",
    Path.cwd() / "wealth_engine.py",
    Path.cwd() / "Wealth Engine.py",
    Path.cwd() / "script.py",
]
PLANS_DIR = BASE_DIR / "plans"
REPORTS_DIR = BASE_DIR / "reports"
RESULTS_DIR = BASE_DIR / "results"
PACKAGES_DIR = BASE_DIR / "packages"
EXPORTS_DIR = BASE_DIR / "exports"
STATE_DIR = BASE_DIR / "state"
ARCHIVE_DIR = BASE_DIR / "archive"
WATCHLIST_FILE = STATE_DIR / "watchlist.json"
LAST_INPUTS_FILE = STATE_DIR / "last_inputs.json"
JOB_FILE = STATE_DIR / "current_job.json"
JOB_LOG_FILE = STATE_DIR / "current_job.log"
JOB_RUNNER_FILE = STATE_DIR / "background_job_runner.py"
for d in (PLANS_DIR, REPORTS_DIR, RESULTS_DIR, PACKAGES_DIR, EXPORTS_DIR, STATE_DIR, ARCHIVE_DIR):
    d.mkdir(parents=True, exist_ok=True)

for k, v in {
    "last_analysis": None,
    "last_signal": None,
    "last_log": "",
    "last_saved_plan": None,
    "last_package": None,
    "current_job_process": None,
    "last_consumed_job_id": None,
}.items():
    st.session_state.setdefault(k, v)

st.markdown(
    """
<style>
:root { --card-bg: rgba(17, 24, 39, 0.72); --card-border: rgba(148, 163, 184, 0.18); }
.main .block-container { padding-top: 1.65rem; padding-bottom: 2.5rem; max-width: 1320px; }
.wealth-title { font-size: 2.05rem; font-weight: 850; color: #f8fafc; margin-bottom: 0.20rem; letter-spacing: -0.025em; }
.wealth-subtitle { color: #b8c4d6; margin-bottom: 1.35rem; max-width: 820px; }
.decision-good { color: #10b981; font-weight: 850; font-size: 1.05rem; }
.decision-warn { color: #f59e0b; font-weight: 850; font-size: 1.05rem; }
.decision-bad { color: #ef4444; font-weight: 850; font-size: 1.05rem; }
.decision-wait { color: #94a3b8; font-weight: 850; font-size: 1.05rem; }
.small-note { color: #94a3b8; font-size: 0.9rem; }
.nav-section { display: block; clear: both; margin-top: 0.56rem; margin-bottom: 0.16rem; padding-top: 0.10rem; color: #94a3b8; font-size: 0.66rem; font-weight: 850; letter-spacing: 0.12em; text-transform: uppercase; line-height: 1.15; }
.clean-success { border-left: 4px solid #10b981; background: rgba(16,185,129,0.11); padding: 0.8rem 1rem; border-radius: 0.7rem; color: #d1fae5; font-weight: 750; }
.danger-zone { border: 1px solid rgba(239,68,68,0.35); border-radius: 0.8rem; padding: 1rem; background: rgba(127,29,29,0.14); }
div[data-testid="stMetric"] { background: var(--card-bg); border: 1px solid var(--card-border); border-radius: 0.85rem; padding: 0.7rem 0.9rem; }
.stDataFrame { border: 1px solid var(--card-border); border-radius: 0.75rem; overflow: hidden; }
section[data-testid="stSidebar"] .stButton > button {
    width: 100%;
    justify-content: flex-start;
    border-radius: 0.55rem;
    padding: 0.24rem 0.55rem;
    min-height: 1.92rem;
    font-size: 0.88rem;
    font-weight: 700;
}
.sidebar-summary {
    border: 1px solid rgba(148,163,184,0.18);
    background: rgba(15,23,42,0.55);
    border-radius: 0.70rem;
    padding: 0.52rem 0.64rem;
    margin-bottom: 0.58rem;
    color: #cbd5e1;
    line-height: 1.20;
    font-size: 0.80rem;
}
.sidebar-summary strong { color: #f8fafc; }
.download-row { display: flex; gap: 0.75rem; align-items: center; flex-wrap: wrap; margin: 0.35rem 0 0.75rem 0; }
section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] { gap: 0.18rem; }
section[data-testid="stSidebar"] h2 { font-size: 1.18rem; margin-bottom: 0.40rem; }
section[data-testid="stSidebar"] hr { margin: 0.52rem 0; }
section[data-testid="stSidebar"] .stCaptionContainer { font-size: 0.78rem; }
.control-row-spacer { height: 0.35rem; }

section[data-testid="stSidebar"] .sidebar-summary p { margin: 0; }
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] { margin-bottom: 0; }

.section-card {
    border: 1px solid rgba(148,163,184,0.22);
    border-radius: 0.70rem;
    padding: 0.70rem 0.85rem;
    background: rgba(15,23,42,0.42);
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 0.6rem;
    margin-bottom: 0.55rem;
}
.muted-label {
    display: block;
    color: #94a3b8;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 0.15rem;
}

</style>
""",
    unsafe_allow_html=True,
)


def find_core_script() -> Optional[Path]:
    for p in CORE_SCRIPT_CANDIDATES:
        if p.exists():
            return p
    return None


def safe_load_json(path: Path) -> Dict[str, Any]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def safe_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        if x is None or x == "":
            return default
        return float(x)
    except Exception:
        return default


def get_first(d: Dict[str, Any], keys: List[str], default: Any = "") -> Any:
    for k in keys:
        if d.get(k) not in (None, ""):
            return d.get(k)
    return default


def fmt_signed_pct(x: Any) -> str:
    v = safe_float(x)
    return "-" if v is None else f"{v:+.2f}%"


def fmt_abs_pct(x: Any) -> str:
    v = safe_float(x)
    return "-" if v is None else f"{abs(v):.2f}%"


def entry_based_trade_percents(sig: Dict[str, Any]) -> Tuple[Any, Any]:
    """Return risk/reward percentages measured from entry price.

    Old saved signal files may contain risk/reward measured from current price.
    Recalculate in the UI when entry/stop/target are available so the display is correct
    even before the next signal refresh.
    """
    entry = safe_float(sig.get("last_entry_price"))
    stop = safe_float(sig.get("current_stop"))
    target = safe_float(sig.get("current_target"))

    risk = current_distance_to_stop_pct(sig)
    reward = current_distance_to_target_pct(sig)

    if entry is not None and entry > 0:
        if stop is not None:
            risk = ((stop / entry) - 1.0) * 100.0
        if target is not None:
            reward = ((target / entry) - 1.0) * 100.0

    return risk, reward


def current_distance_to_stop_pct(sig: Dict[str, Any]) -> Any:
    current = safe_float(sig.get("current_price"))
    stop = safe_float(sig.get("current_stop"))
    if current is not None and current > 0 and stop is not None:
        return ((stop / current) - 1.0) * 100.0
    return sig.get("distance_to_stop_pct", sig.get("risk_to_stop_pct"))


def current_distance_to_target_pct(sig: Dict[str, Any]) -> Any:
    current = safe_float(sig.get("current_price"))
    target = safe_float(sig.get("current_target"))
    if current is not None and current > 0 and target is not None:
        return ((target / current) - 1.0) * 100.0
    return sig.get("distance_to_target_pct", sig.get("reward_to_target_pct"))


def fmt_num(x: Any, digits: int = 2) -> str:
    v = safe_float(x)
    return "-" if v is None else f"{v:,.{digits}f}"


def symbol_from_json(path: Path, payload: Optional[Dict[str, Any]] = None) -> str:
    payload = payload or safe_load_json(path)
    summary = payload.get("summary", {})
    symbol = summary.get("symbol") or summary.get("original_analysis_symbol")
    if symbol:
        return str(symbol).upper()
    name = path.name
    for token in ["_best_strategy_", "_signal_", "_replay_", "_plan"]:
        if token in name:
            return name.split(token)[0].upper()
    return path.stem.split("_")[0].upper()


def timeframe_from_payload(payload: Dict[str, Any]) -> str:
    summary = payload.get("summary", {})
    params = payload.get("params", {})
    return str(summary.get("timeframe") or summary.get("best_timeframe") or summary.get("replay_timeframe") or params.get("timeframe") or "-")


def decision_class(decision: str) -> str:
    d = str(decision or "")
    if "PROFIT" in d or "NEAR_TARGET" in d:
        return "decision-good"
    if "NEAR_STOP" in d:
        return "decision-bad"
    if "LOSING" in d:
        return "decision-warn"
    return "decision-wait"


def signal_short_reason(sig: Dict[str, Any]) -> str:
    status = str(sig.get("current_status") or "").upper()
    action = str(sig.get("action") or "").upper()
    if status == "OPEN":
        if "LOSING" in action:
            return "Open trade. Price is below entry but still above the stop."
        if "PROFIT" in action:
            return "Open trade. Position is currently profitable."
        if "NEAR_STOP" in action:
            return "Open trade. Price is close to the stop."
        return "Open trade. Follow the active stop and target."
    if status == "ENTRY_PENDING_NEXT_BAR":
        return "Entry signal detected. Watch the next bar open."
    if "WAIT_AFTER_EXIT" in action:
        return "Last trade is closed. Waiting for a new setup."
    if "WAIT" in action:
        return "No active setup right now."
    return "Current strategy snapshot."


def next_serial(symbol: str, mode: str, ext: str, folder: Path = REPORTS_DIR) -> int:
    pattern = f"{symbol.upper()}_{mode.upper()}_*.{ext.lstrip('.')}"
    max_n = 0
    for p in folder.glob(pattern):
        m = re.search(rf"{re.escape(symbol.upper())}_{re.escape(mode.upper())}_(\d+)\.{re.escape(ext.lstrip('.'))}$", p.name)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return max_n + 1


def standard_report_name(symbol: str, mode: str, ext: str) -> Path:
    # Official saved report name. Overwrites instead of creating many versions.
    return REPORTS_DIR / f"{symbol.upper()}_{mode.upper()}.{ext.lstrip('.')}"

def standard_package_name(symbol: str, mode: str = "OPTIMIZE", force_new: bool = False) -> Path:
    # Official saved package name. Overwrites instead of creating many versions.
    symbol = symbol.upper()
    mode = mode.upper()
    if not force_new:
        return PACKAGES_DIR / f"{symbol}_{mode}.zip"
    n = next_serial(symbol, mode, "zip", PACKAGES_DIR)
    return PACKAGES_DIR / f"{symbol}_{mode}_{n:03d}.zip"

def latest_matching(folder: Path, patterns: List[str], since_ts: Optional[float] = None) -> Optional[Path]:
    files: List[Path] = []
    for pat in patterns:
        files.extend(folder.glob(pat))
    if since_ts is not None:
        files = [p for p in files if p.stat().st_mtime >= since_ts]
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


def latest_report(symbol: str, mode: str, ext: str) -> Optional[Path]:
    symbol = symbol.upper()
    mode = mode.upper()
    ext = ext.lstrip(".")

    active = REPORTS_DIR / f"{symbol}_{mode}.{ext}"
    if active.exists():
        return active

    files = list(REPORTS_DIR.glob(f"{symbol}_{mode}_*.{ext}"))
    return max(files, key=lambda p: p.stat().st_mtime) if files else None

def latest_optimize_json(symbol: str) -> Optional[Path]:
    return latest_report(symbol, "OPTIMIZE", "json")


def active_param_file(symbol: str) -> Optional[Path]:
    # Only saved active plans are official inputs for Dashboard / Signal / Reports.
    symbol = symbol.upper()
    plan = PLANS_DIR / f"{symbol}_plan.json"
    return plan if plan.exists() else None


def active_param_files() -> List[Path]:
    # Dashboard shows saved active plans only.
    return sorted(PLANS_DIR.glob("*_plan.json"))


def active_plan_files() -> List[Path]:
    return sorted(PLANS_DIR.glob("*_plan.json"))


def active_plan_symbols() -> List[str]:
    return sorted([p.name.replace("_plan.json", "").upper() for p in active_plan_files()])

def persist_last_analysis(data: Dict[str, Any]) -> None:
    """Save latest analysis to disk so changing pages or browser reconnects does not lose it."""
    try:
        out = STATE_DIR / "last_analysis.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
    except Exception:
        pass


def load_last_analysis() -> Optional[Dict[str, Any]]:
    # Load latest analysis from memory or state file. This is temporary until Save as Active Plan.
    current = st.session_state.get("last_analysis")
    if current:
        return current

    state_file = STATE_DIR / "last_analysis.json"
    if state_file.exists():
        data = safe_load_json(state_file)
        if data:
            st.session_state["last_analysis"] = data
            return data

    return None

def guide_rows() -> List[Tuple[str, str, str]]:
    return [
        ("Compound Profit %", "Meaning", "Compounded return based on the selected investment amount."),
        ("Strategy P/L", "Meaning", "Backtest result based on the strategy settings."),
        ("Max Drawdown", "Meaning", "Largest equity decline from peak to trough."),
        ("Win Rate", "Meaning", "Winning trades divided by total closed trades."),
        ("Profit Factor", "Meaning", "Gross profit divided by gross loss."),
        ("Sharpe", "Meaning", "Return quality adjusted for volatility."),
        ("Has Open Trade", "Meaning", "YES means the plan currently has an open position."),
        ("Smart Decision", "Meaning", "Current action label for the plan."),
        ("Current Stop", "Meaning", "Current protective stop level."),
        ("Current Target", "Meaning", "Current profit target."),
        ("Risk to Stop %", "Meaning", "Stop distance from entry price."),
        ("Reward to Target %", "Meaning", "Target distance from entry price."),
        ("Latest Trades", "Meaning", "Most recent closed/open trades available from Signal or Replay."),
    ]

def ensure_excel_guide_sheet(path: Path) -> None:
    if load_workbook is None or not path.exists() or path.suffix.lower() != ".xlsx":
        return
    try:
        wb = load_workbook(path)
        if "Guide" not in wb.sheetnames:
            ws = wb.create_sheet("Guide")
            ws.append(["Field / Column", "Meaning", "Description"])
            for row in guide_rows():
                ws.append(list(row))
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF4EC")
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 70
        wb.save(path)
    except Exception:
        return


def rename_latest_outputs(symbol: str, mode: str, since_ts: float) -> Dict[str, Optional[Path]]:
    symbol = symbol.upper()
    mode = mode.upper()
    mode_map = {
        "OPTIMIZE": [f"{symbol}_best_strategy_*.xlsx", f"{symbol}_best_strategy_*.json"],
        "REPLAY": [f"{symbol}_replay_*.xlsx", f"{symbol}_replay_*.json"],
        "SIGNAL": [f"{symbol}_signal_*.xlsx", f"{symbol}_signal_*.json"],
        "DASHBOARD": ["strategy_dashboard_*.xlsx", "strategy_dashboard_*.json"],
    }
    patterns = mode_map.get(mode, [])
    out: Dict[str, Optional[Path]] = {"xlsx": None, "json": None}
    for ext in ("xlsx", "json"):
        candidates = [p for p in patterns if p.endswith(f".{ext}")]
        src = latest_matching(RESULTS_DIR, candidates, since_ts=since_ts)
        if src and src.exists():
            dst = standard_report_name(symbol if mode != "DASHBOARD" else "ALL", mode, ext)
            shutil.copy2(src, dst)
            if ext == "xlsx":
                ensure_excel_guide_sheet(dst)
            out[ext] = dst
    return out


def import_latest_output(symbol: str, mode: str, ext: str) -> Optional[Path]:
    """
    Robust fallback:
    If the file was created under /results but did not get copied to /reports,
    import the latest raw output and rename it to the clean report pattern.
    """
    symbol = symbol.upper()
    mode = mode.upper()
    ext = ext.lstrip(".")

    patterns = {
        "OPTIMIZE": [f"{symbol}_best_strategy_*.{ext}", f"{symbol}_OPTIMIZE_*.{ext}"],
        "SIGNAL": [f"{symbol}_signal_*.{ext}", f"{symbol}_SIGNAL_*.{ext}"],
        "REPLAY": [f"{symbol}_replay_*.{ext}", f"{symbol}_REPLAY_*.{ext}"],
    }.get(mode, [f"{symbol}_{mode}_*.{ext}"])

    raw = latest_matching(RESULTS_DIR, patterns, since_ts=None)
    if raw and raw.exists():
        dst = standard_report_name(symbol, mode, ext)
        shutil.copy2(raw, dst)
        if ext == "xlsx":
            ensure_excel_guide_sheet(dst)
        return dst

    return latest_report(symbol, mode, ext)


def ensure_report_available(symbol: str, mode: str) -> Dict[str, Optional[Path]]:
    """
    Return latest clean report files. If missing, import them from /results.
    """
    symbol = symbol.upper()
    mode = mode.upper()
    xlsx = latest_report(symbol, mode, "xlsx") or import_latest_output(symbol, mode, "xlsx")
    jsn = latest_report(symbol, mode, "json") or import_latest_output(symbol, mode, "json")
    return {"xlsx": xlsx, "json": jsn}


def save_active_plan(symbol: str, json_path: Path) -> Path:
    dst = PLANS_DIR / f"{symbol.upper()}_plan.json"
    shutil.copy2(json_path, dst)
    st.session_state["last_saved_plan"] = str(dst)
    return dst


def list_plan_files() -> List[Path]:
    return sorted(PLANS_DIR.glob("*_plan.json"))


def list_symbols_from_plans() -> List[str]:
    return sorted([symbol_from_json(p) for p in list_plan_files()])


def list_symbols_from_reports_and_plans() -> List[str]:
    symbols = set(list_symbols_from_plans())
    for p in REPORTS_DIR.glob("*_*_*.xlsx"):
        symbols.add(p.name.split("_")[0].upper())
    for p in REPORTS_DIR.glob("*_*_*.json"):
        symbols.add(p.name.split("_")[0].upper())
    return sorted(symbols)


def newest_raw_output(symbol: str, raw_mode: str, ext: str, since_ts: Optional[float] = None) -> Optional[Path]:
    symbol = symbol.upper()
    ext = ext.lstrip(".")
    files = list(RESULTS_DIR.glob(f"{symbol}_{raw_mode}_*.{ext}"))
    if since_ts is not None:
        files = [p for p in files if p.stat().st_mtime >= since_ts]
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


def save_active_report(symbol: str, mode: str, source: Optional[Path]) -> Optional[Path]:
    if not source or not Path(source).exists():
        return None
    symbol = symbol.upper()
    mode = mode.upper()
    source = Path(source)
    dest = REPORTS_DIR / f"{symbol}_{mode}{source.suffix.lower()}"
    try:
        if source.suffix.lower() == ".xlsx":
            ensure_excel_guide_sheet(source)
        shutil.copy2(source, dest)
        if dest.suffix.lower() == ".xlsx":
            ensure_excel_guide_sheet(dest)
        return dest
    except Exception as exc:
        st.warning(f"Could not save {source.name} to reports/: {exc}")
        return None


def save_active_outputs_from_last(symbol: str, analysis_data: Dict[str, Any]) -> Dict[str, Dict[str, Optional[Path]]]:
    # Save official reports only when Save as Active Plan is clicked.
    symbol = symbol.upper()
    saved = {
        "OPTIMIZE": {"xlsx": None, "json": None},
        "SIGNAL": {"xlsx": None, "json": None},
    }

    opt_json = Path(analysis_data.get("opt_json") or "")
    opt_xlsx = Path(analysis_data.get("opt_xlsx") or "")
    sig_json = Path(analysis_data.get("signal_json") or "")
    sig_xlsx = Path(analysis_data.get("signal_xlsx") or "")

    if not opt_json.exists():
        opt_json = newest_raw_output(symbol, "best_strategy", "json") or opt_json
    if not opt_xlsx.exists():
        opt_xlsx = newest_raw_output(symbol, "best_strategy", "xlsx") or opt_xlsx
    if not sig_json.exists():
        sig_json = newest_raw_output(symbol, "signal", "json") or sig_json
    if not sig_xlsx.exists():
        sig_xlsx = newest_raw_output(symbol, "signal", "xlsx") or sig_xlsx

    saved["OPTIMIZE"]["json"] = save_active_report(symbol, "OPTIMIZE", opt_json)
    saved["OPTIMIZE"]["xlsx"] = save_active_report(symbol, "OPTIMIZE", opt_xlsx)
    saved["SIGNAL"]["json"] = save_active_report(symbol, "SIGNAL", sig_json)
    saved["SIGNAL"]["xlsx"] = save_active_report(symbol, "SIGNAL", sig_xlsx)

    return saved


def create_symbol_package(symbol: str, package_mode: str = "OPTIMIZE", force_new: bool = False) -> Tuple[Path, Dict[str, bool]]:
    symbol = symbol.upper()
    package_path = standard_package_name(symbol, package_mode, force_new=force_new)
    availability = {"plan": False, "optimize": False, "signal": False, "replay": False}
    plan = PLANS_DIR / f"{symbol}_plan.json"
    files = {
        f"{symbol}_PLAN.json": plan if plan.exists() else None,
        f"{symbol}_OPTIMIZE.xlsx": latest_report(symbol, "OPTIMIZE", "xlsx"),
        f"{symbol}_OPTIMIZE.json": latest_report(symbol, "OPTIMIZE", "json"),
        f"{symbol}_SIGNAL.xlsx": latest_report(symbol, "SIGNAL", "xlsx"),
        f"{symbol}_SIGNAL.json": latest_report(symbol, "SIGNAL", "json"),
        f"{symbol}_REPLAY.xlsx": latest_report(symbol, "REPLAY", "xlsx"),
        f"{symbol}_REPLAY.json": latest_report(symbol, "REPLAY", "json"),
    }
    with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for arcname, src in files.items():
            if src and src.exists():
                if src.suffix.lower() == ".xlsx":
                    ensure_excel_guide_sheet(src)
                zipf.write(src, arcname=arcname)
                if "PLAN" in arcname:
                    availability["plan"] = True
                elif "OPTIMIZE" in arcname:
                    availability["optimize"] = True
                elif "SIGNAL" in arcname:
                    availability["signal"] = True
                elif "REPLAY" in arcname:
                    availability["replay"] = True
    st.session_state["last_package"] = str(package_path)
    return package_path, availability


def latest_package(symbol: str) -> Optional[Path]:
    symbol = symbol.upper()
    active = PACKAGES_DIR / f"{symbol}_OPTIMIZE.zip"
    if active.exists():
        return active
    files = list(PACKAGES_DIR.glob(f"{symbol}_OPTIMIZE_*.zip"))
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


@st.cache_data(show_spinner=False)
def cached_file_bytes(path_str: str, size: int, mtime_ns: int) -> bytes:
    return Path(path_str).read_bytes()


def download_button_for_file(path: Path, label: Optional[str] = None, key_prefix: str = "download") -> None:
    """Stable download button for Excel / JSON / ZIP on local Streamlit."""
    path = Path(path)

    if not path.exists() or not path.is_file():
        st.warning(f"File not found: {path.name or path}")
        return

    stat = path.stat()
    data = cached_file_bytes(str(path), stat.st_size, stat.st_mtime_ns)

    suffix = path.suffix.lower()
    label = label or f"Download {path.name}"

    # ZIP fix: avoid Streamlit temporary /media link
    if suffix == ".zip":
        b64 = base64.b64encode(data).decode("utf-8")
        href = f'''
        <a download="{path.name}"
           href="data:application/zip;base64,{b64}"
           style="
               display:inline-block;
               padding:0.55rem 0.85rem;
               border:1px solid rgba(250,250,250,0.25);
               border-radius:0.5rem;
               color:white;
               text-decoration:none;
               font-weight:600;
           ">
           {label}
        </a>
        '''
        st.markdown(href, unsafe_allow_html=True)
        return

    if suffix == ".xlsx":
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif suffix == ".json":
        mime = "application/json"
    else:
        mime = "application/octet-stream"

    st.download_button(
        label=label,
        data=data,
        file_name=path.name,
        mime=mime,
        key=f"{key_prefix}_{path.name}_{stat.st_size}_{stat.st_mtime_ns}",
    )

def download_dataframe_csv(df: pd.DataFrame, filename: str, label: str, key: str) -> None:
    csv_data = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label=label,
        data=csv_data,
        file_name=filename,
        mime="text/csv",
        key=key,
    )


def run_core_command_stream(args: List[str], log_title: str, trials: Optional[int] = None) -> Tuple[int, str]:
    core = find_core_script()
    if core is None:
        return 1, "Core script not found. Expected: Wealth Engine.py"
    cmd = [sys.executable, "-u", str(core)] + args
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    progress = st.progress(0, text="Starting Wealth Engine...")
    status_box = st.empty()
    log_lines: List[str] = []
    with st.expander(log_title, expanded=False):
        log_area = st.empty()
    start = time.time()
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
        )
        assert proc.stdout is not None
        last_trial = 0
        for line in proc.stdout:
            line = line.rstrip("\n")
            log_lines.append(line)
            m = re.search(r"\[\s*(\d+)\]", line)
            if m and trials:
                last_trial = max(last_trial, int(m.group(1)))
                progress.progress(min(0.95, max(0.02, last_trial / max(trials, 1))), text=f"Optimizing... {last_trial}/{trials} trials")
            elif "Loading data" in line:
                progress.progress(0.05, text="Loading data...")
            elif "Starting Optimization" in line:
                progress.progress(0.08, text="Starting optimization...")
            elif "Best Result" in line:
                progress.progress(0.97, text="Preparing report...")
            elif "Saved files" in line:
                progress.progress(0.99, text="Saving files...")
            if line.strip():
                status_box.caption(line[-220:])
            if len(log_lines) % 10 == 0:
                log_area.code("\n".join(log_lines[-120:]))
        rc = proc.wait()
        elapsed = time.time() - start
        progress.progress(1.0, text=f"Done in {elapsed:.1f}s" if rc == 0 else "Failed")
        log_area.code("\n".join(log_lines[-250:]))
        return rc, "\n".join(log_lines)
    except Exception as e:
        return 1, f"Execution error: {e}"



def plan_run_args(plan_file: Path, fallback_years: int = 1, latest_end: bool = False) -> List[str]:
    payload = safe_load_json(plan_file)
    summary = payload.get("summary", {}) if payload else {}
    args: List[str] = []

    start_date = summary.get("analysis_start")
    end_date = summary.get("analysis_end")
    investment_amount = summary.get("investment_amount") or summary.get("initial_capital") or 100.0

    if start_date:
        args.extend(["--start-date", str(start_date)])
        # For Signal / Auto Refresh, never reuse the old Optimize end date.
        # Missing --end-date makes wealth_engine.py fetch up to latest available.
        if end_date and not latest_end:
            args.extend(["--end-date", str(end_date)])
    else:
        args.extend(["--years", str(fallback_years)])

    args.extend(["--investment-amount", str(float(investment_amount))])
    return args

def run_signal_for_plan(plan_file: Path, years: int = 1, adjust: str = "splits", pause_sec: Optional[float] = None) -> Dict[str, Optional[Path]]:
    symbol = symbol_from_json(plan_file)
    start_ts = datetime.now().timestamp()
    extra_args: List[str] = []
    if pause_sec is not None:
        extra_args.extend(["--pause-sec", str(max(float(pause_sec), 0.0))])
    rc, log = run_core_command_stream([
        symbol, "--mode", "signal",
        *plan_run_args(plan_file, fallback_years=years, latest_end=True),
        "--adjust", adjust,
        *extra_args,
        "--params-file", str(plan_file), "--output", str(RESULTS_DIR)
    ], log_title=f"Signal Log - {symbol}")
    st.session_state["last_log"] = log
    if rc != 0:
        st.error(f"Signal failed for {symbol}.")
        return {"xlsx": None, "json": None}
    outputs = rename_latest_outputs(symbol, "SIGNAL", start_ts)
    if not outputs.get("json"):
        outputs = ensure_report_available(symbol, "SIGNAL")
    create_symbol_package(symbol, "OPTIMIZE")
    return outputs


def wait_seconds_from_api_log(log: str, api_limit_per_minute: int = 8, buffer_seconds: int = 10, fallback_seconds: int = 75) -> int:
    """Estimate safe wait time after a Twelve Data refresh.

    If Twelve Data reports used credits, wait enough full minutes for the free quota window.
    If no credit count is present, use a conservative fallback.
    """
    limit = max(int(api_limit_per_minute or 8), 1)
    buffer = max(int(buffer_seconds or 0), 0)
    fallback = max(int(fallback_seconds or 0), buffer)
    matches = re.findall(r"(\d+)\s+API credits?\s+were used", str(log), flags=re.IGNORECASE)
    if not matches:
        return fallback
    used = max(int(x) for x in matches)
    blocks = max(1, math.ceil(used / limit))
    return int(blocks * 60 + buffer)


def countdown_wait(seconds: int, label: str = "Waiting for API window") -> None:
    seconds = int(max(seconds, 0))
    if seconds <= 0:
        return
    box = st.empty()
    bar = st.progress(0, text=f"{label}: {seconds}s")
    for remaining in range(seconds, 0, -1):
        done = (seconds - remaining) / max(seconds, 1)
        bar.progress(min(done, 1.0), text=f"{label}: {remaining}s")
        if remaining == seconds or remaining % 10 == 0 or remaining <= 5:
            box.caption(f"Next symbol starts in {remaining}s")
        time.sleep(1)
    bar.progress(1.0, text="Continuing...")
    box.empty()


def run_replay_for_plan(plan_file: Path, years: int = 1, adjust: str = "splits") -> Dict[str, Optional[Path]]:
    symbol = symbol_from_json(plan_file)
    start_ts = datetime.now().timestamp()
    rc, log = run_core_command_stream([
        symbol, "--mode", "replay",
        # Replay should rebuild the full trade history up to the latest available data.
        # It must not overwrite OPTIMIZE; it only refreshes REPLAY reports.
        *plan_run_args(plan_file, fallback_years=years, latest_end=True),
        "--adjust", adjust,
        "--params-file", str(plan_file), "--output", str(RESULTS_DIR)
    ], log_title=f"Replay Log - {symbol}")
    st.session_state["last_log"] = log
    if rc != 0:
        st.error(f"Replay failed for {symbol}.")
        return {"xlsx": None, "json": None}
    outputs = rename_latest_outputs(symbol, "REPLAY", start_ts)
    if not outputs.get("json"):
        outputs = ensure_report_available(symbol, "REPLAY")
    create_symbol_package(symbol, "OPTIMIZE")
    return outputs


def extract_signal_payload(symbol: str) -> Optional[Dict[str, Any]]:
    # Check clean reports first; if missing, import the latest raw signal from /results.
    sig_json = latest_report(symbol, "SIGNAL", "json")
    if sig_json is None:
        sig_json = import_latest_output(symbol, "SIGNAL", "json")
    return safe_load_json(sig_json) if sig_json else None


def extract_replay_payload(symbol: str) -> Optional[Dict[str, Any]]:
    # Replay is separate from Optimize and Signal. Use it for trade history / recent events.
    replay_json = latest_report(symbol, "REPLAY", "json")
    if replay_json is None:
        replay_json = import_latest_output(symbol, "REPLAY", "json")
    return safe_load_json(replay_json) if replay_json else None


def latest_replay_files(symbol: str) -> Dict[str, Optional[Path]]:
    return ensure_report_available(symbol, "REPLAY")


@st.cache_data(show_spinner=False)
def cached_excel_sheet(path_str: str, sheet_name: str, size: int, mtime_ns: int) -> pd.DataFrame:
    try:
        return pd.read_excel(path_str, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def load_report_trades_df(symbol: str, modes: Optional[List[str]] = None, min_rows: int = 4) -> Tuple[pd.DataFrame, str]:
    """Load the best available Trades sheet for a symbol.

    Preference is REPLAY -> SIGNAL -> OPTIMIZE, but for the Signals page we
    prefer a source with at least min_rows so Latest 4 Trades can actually show
    four rows when the full Optimize history has them.
    """
    modes = modes or ["REPLAY", "SIGNAL", "OPTIMIZE"]
    candidates: List[Tuple[str, pd.DataFrame]] = []
    for mode in modes:
        xlsx = latest_report(symbol, mode, "xlsx")
        if xlsx is None:
            xlsx = import_latest_output(symbol, mode, "xlsx")
        if xlsx and xlsx.exists():
            stat = xlsx.stat()
            df = cached_excel_sheet(str(xlsx), "Trades", stat.st_size, stat.st_mtime_ns)
            if not df.empty:
                candidates.append((mode, df))
                if len(df) >= min_rows:
                    return df, mode
    if candidates:
        return max(candidates, key=lambda item: len(item[1]))
    return pd.DataFrame(), "-"


def load_replay_trades_df(symbol: str) -> pd.DataFrame:
    df, _ = load_report_trades_df(symbol, ["REPLAY"])
    return df


def summary_compound_pct(summary: Dict[str, Any]) -> Any:
    return get_first(summary, ["compound_net_profit_pct", "net_profit_pct", "compound_profit_pct"])


def summary_strategy_pl(summary: Dict[str, Any]) -> Any:
    return get_first(summary, ["strategy_net_profit_pct", "strategy_p_l_pct", "strategy_pl_pct", "strategy_profit_pct", "strategy_p_l"])



def dashboard_row(plan_file: Path) -> Dict[str, Any]:
    plan_payload = safe_load_json(plan_file)
    plan_summary = plan_payload.get("summary", {})
    symbol = symbol_from_json(plan_file, plan_payload)
    signal_payload = extract_signal_payload(symbol)
    sig = signal_payload.get("current_signal", {}) if signal_payload else {}
    sig_summary = signal_payload.get("summary", {}) if signal_payload else {}
    summary = {**plan_summary, **sig_summary}
    has_open = sig.get("current_status") == "OPEN" if sig else bool(summary.get("has_open_trade"))
    decision = sig.get("action") or ("OPEN" if has_open else "WAIT")
    return {
        "Symbol": symbol,
        "Period": summary.get("analysis_period") or f"{summary.get('analysis_start', '')} -> {summary.get('analysis_end', '')}",
        "Timeframe": timeframe_from_payload(signal_payload or plan_payload),
        "Has Open Trade": "YES" if has_open else "NO",
        "Smart Decision": decision,
        "Investment Amount": "$" + fmt_num(summary.get("investment_amount", summary.get("initial_capital")), 2),
        "Final Capital": "$" + fmt_num(summary.get("final_capital"), 2),
        "Compound Profit %": fmt_signed_pct(summary_compound_pct(summary)),
        "Strategy P/L": fmt_signed_pct(summary_strategy_pl(summary)),
        "Trades": get_first(summary, ["num_trades", "trades"], "-"),
        "Win Rate": fmt_abs_pct(get_first(summary, ["win_rate"], None)),
        "Max Drawdown": fmt_abs_pct(get_first(summary, ["max_drawdown_pct"], None)),
    }



def decision_color_css(value: Any) -> str:
    text = str(value).upper()
    if any(x in text for x in ["BUY", "HOLD", "OPEN", "GOOD"]):
        return "color: #10b981; font-weight: 800;"
    if any(x in text for x in ["NEAR_STOP", "WARN", "CAUTION"]):
        return "color: #f59e0b; font-weight: 800;"
    if any(x in text for x in ["EXIT", "STOP", "BAD", "SELL"]):
        return "color: #ef4444; font-weight: 800;"
    if "WAIT" in text:
        return "color: #94a3b8; font-weight: 800;"
    return ""


def yes_no_color_css(value: Any) -> str:
    text = str(value).upper()
    if text == "YES":
        return "color: #10b981; font-weight: 800;"
    if text == "NO":
        return "color: #94a3b8; font-weight: 800;"
    return ""


def pct_color_css(value: Any) -> str:
    text = str(value).replace("%", "").replace("+", "").strip()
    try:
        num = float(text)
    except Exception:
        return ""
    if num > 0:
        return "color: #10b981; font-weight: 700;"
    if num < 0:
        return "color: #ef4444; font-weight: 700;"
    return ""


def style_dashboard_df(df: pd.DataFrame):
    if df.empty:
        return df
    styler = df.style
    if "Smart Decision" in df.columns:
        styler = styler.map(decision_color_css, subset=["Smart Decision"])
    if "Open Trade" in df.columns:
        styler = styler.map(yes_no_color_css, subset=["Open Trade"])
    if "Has Open Trade" in df.columns:
        styler = styler.map(yes_no_color_css, subset=["Has Open Trade"])
    for col in ["Compound Profit %", "Strategy P/L", "Current PnL %"]:
        if col in df.columns:
            styler = styler.map(pct_color_css, subset=[col])
    return styler


def official_symbol_files(symbol: str) -> List[Path]:
    """All official files for an active saved symbol."""
    symbol = symbol.upper()
    candidates = [
        PLANS_DIR / f"{symbol}_plan.json",
        PACKAGES_DIR / f"{symbol}_OPTIMIZE.zip",
    ]
    candidates.extend(sorted(REPORTS_DIR.glob(f"{symbol}_OPTIMIZE.*")))
    candidates.extend(sorted(REPORTS_DIR.glob(f"{symbol}_SIGNAL.*")))
    candidates.extend(sorted(REPORTS_DIR.glob(f"{symbol}_REPLAY.*")))
    return [p for p in candidates if p.exists()]


def archive_active_symbol(symbol: str) -> Path:
    """Move official files to archive instead of deleting forever."""
    symbol = symbol.upper()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest_dir = ARCHIVE_DIR / f"{symbol}_{stamp}"
    dest_dir.mkdir(parents=True, exist_ok=True)

    moved = []
    for src in official_symbol_files(symbol):
        target = dest_dir / src.name
        try:
            shutil.move(str(src), str(target))
            moved.append(target)
        except Exception as exc:
            st.warning(f"Could not archive {src.name}: {exc}")

    # Remove cached package pointer if it pointed to archived file.
    st.session_state["last_package"] = None
    return dest_dir



def _load_symbol_bucket(path: Path) -> List[str]:
    if not path.exists():
        return []
    data = safe_load_json(path)
    items = data.get("symbols", []) if isinstance(data, dict) else []
    return sorted({str(x).strip().upper() for x in items if str(x).strip()})


def _save_symbol_bucket(path: Path, symbols: List[str]) -> None:
    clean = sorted({str(s).strip().upper() for s in symbols if str(s).strip()})
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"symbols": clean, "updated_at": datetime.now().isoformat(timespec="seconds")}, f, indent=2)


def load_watchlist() -> List[str]:
    return _load_symbol_bucket(WATCHLIST_FILE)


def save_watchlist(symbols: List[str]) -> None:
    _save_symbol_bucket(WATCHLIST_FILE, symbols)


def load_my_stocks() -> List[str]:
    active = set(active_plan_symbols())
    watch = set(load_watchlist())
    return sorted(active - watch)


def save_my_stocks(symbols: List[str]) -> None:
    # My Stock is the default bucket: active plans that are not in watchlist.
    return


def add_to_watchlist(symbol: str) -> None:
    symbol = symbol.strip().upper()
    if not symbol:
        return
    watch = set(load_watchlist())
    watch.add(symbol)
    save_watchlist(sorted(watch))


def remove_from_watchlist(symbol: str) -> None:
    symbol = symbol.strip().upper()
    items = [s for s in load_watchlist() if s != symbol]
    save_watchlist(items)


def add_to_my_stocks(symbol: str) -> None:
    symbol = symbol.strip().upper()
    if not symbol:
        return
    remove_from_watchlist(symbol)


def classified_symbols(kind: str) -> List[str]:
    active = set(active_plan_symbols())
    if kind == "watchlist":
        return sorted(active & set(load_watchlist()))
    if kind == "my_stock":
        return sorted(active - set(load_watchlist()))
    return sorted(active)


def symbol_section(symbol: str) -> str:
    symbol = symbol.strip().upper()
    if symbol in set(load_watchlist()):
        return "Watchlist"
    return "My Stock"


def move_symbol_to_section(symbol: str, target: str) -> str:
    symbol = symbol.strip().upper()
    if target == "Watchlist":
        add_to_watchlist(symbol)
        return "Watchlist"
    add_to_my_stocks(symbol)
    return "My Stock"


def plan_files_for_symbols(symbols: List[str]) -> List[Path]:
    out = []
    for symbol in symbols:
        p = PLANS_DIR / f"{symbol.upper()}_plan.json"
        if p.exists():
            out.append(p)
    return sorted(out)


def save_last_to_category(symbol: str, analysis_data: Dict[str, Any], category: str) -> Optional[Path]:
    saved = save_active_outputs_from_last(symbol, analysis_data)
    report_json = saved["OPTIMIZE"].get("json")
    if not report_json:
        st.error("Could not save Optimize JSON to reports/.")
        return None

    plan_path = save_active_plan(symbol, report_json)
    package_path, _ = create_symbol_package(symbol, "OPTIMIZE", force_new=False)

    analysis_data["opt_json"] = str(report_json)
    analysis_data["opt_xlsx"] = str(saved["OPTIMIZE"].get("xlsx") or analysis_data.get("opt_xlsx") or "")
    analysis_data["signal_json"] = str(saved["SIGNAL"].get("json") or analysis_data.get("signal_json") or "")
    analysis_data["signal_xlsx"] = str(saved["SIGNAL"].get("xlsx") or analysis_data.get("signal_xlsx") or "")
    analysis_data["package"] = str(package_path)
    analysis_data["saved_plan"] = str(plan_path)

    if category == "watchlist":
        add_to_watchlist(symbol)
        label = "Watchlist"
    else:
        add_to_my_stocks(symbol)
        label = "My Stock"

    st.session_state["last_analysis"] = analysis_data
    persist_last_analysis(analysis_data)
    st.session_state["last_saved_plan"] = str(plan_path)
    st.success(f"Saved to {label}: {symbol}")
    return plan_path


def run_auto_refresh_if_due(interval_minutes: int) -> None:
    """Dashboard auto refresh: browser reruns page and this runs signals when due."""
    if not st.session_state.get("auto_refresh_enabled"):
        return

    interval_seconds = max(1, int(interval_minutes)) * 60
    now = time.time()
    last_run = float(st.session_state.get("auto_refresh_last_run") or 0)

    if now - last_run >= interval_seconds:
        param_files = active_param_files()
        for p in param_files:
            run_signal_for_plan(p, years=int(st.session_state.get("auto_refresh_years", 1)))
        st.session_state["auto_refresh_last_run"] = now
        st.toast("Auto Refresh completed.")

    # Streamlit-only simple browser refresh. No extra package required.
    st.markdown(
        f"<script>setTimeout(function(){{ window.location.reload(); }}, {interval_seconds * 1000});</script>",
        unsafe_allow_html=True,
    )





def default_last_inputs() -> Dict[str, Any]:
    end_d = date.today()
    start_d = end_d - timedelta(days=365)
    return {
        "ticker": "IREN",
        "start_date": start_d.isoformat(),
        "end_date": end_d.isoformat(),
        "investment_amount": 100.0,
        "timeframe": "4h",
        "trials": 500,
        "jobs": 8,
        "min_trades": 12,
    }


def _date_from_input(value: Any, fallback: date) -> date:
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return fallback


def load_last_inputs() -> Dict[str, Any]:
    data = {}
    if LAST_INPUTS_FILE.exists():
        data = safe_load_json(LAST_INPUTS_FILE)
    defaults = default_last_inputs()
    if isinstance(data, dict):
        # Backward compatibility with old state/last_inputs.json that only had years.
        if "start_date" not in data and "years" in data:
            end_d = date.today()
            start_d = end_d - timedelta(days=365 * int(data.get("years", 1)))
            data["start_date"] = start_d.isoformat()
            data["end_date"] = end_d.isoformat()
        defaults.update({k: data.get(k, defaults[k]) for k in defaults})
    return defaults


def save_last_inputs(inputs: Dict[str, Any]) -> None:
    clean = {
        "ticker": str(inputs.get("ticker", "IREN")).strip().upper() or "IREN",
        "start_date": str(inputs.get("start_date")),
        "end_date": str(inputs.get("end_date")),
        "investment_amount": float(inputs.get("investment_amount", 100.0)),
        "timeframe": str(inputs.get("timeframe", "4h")),
        "trials": int(inputs.get("trials", 500)),
        "jobs": int(inputs.get("jobs", 8)),
        "min_trades": int(inputs.get("min_trades", 12)),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    with open(LAST_INPUTS_FILE, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2)


def current_form_inputs() -> Dict[str, Any]:
    return {
        "ticker": str(st.session_state.get("analyze_ticker", "IREN")).strip().upper(),
        "start_date": str(st.session_state.get("analyze_start_date")),
        "end_date": str(st.session_state.get("analyze_end_date")),
        "investment_amount": float(st.session_state.get("analyze_investment_amount", 100.0)),
        "timeframe": str(st.session_state.get("analyze_timeframe", "4h")),
        "trials": int(st.session_state.get("analyze_trials", 500)),
        "jobs": int(st.session_state.get("analyze_jobs", 8)),
        "min_trades": int(st.session_state.get("analyze_min_trades", 12)),
    }


def inputs_for_display(inputs: Dict[str, Any]) -> str:
    return (
        f"{inputs.get('ticker')} | TF={inputs.get('timeframe')} | "
        f"{inputs.get('start_date')} -> {inputs.get('end_date')} | "
        f"Investment=${float(inputs.get('investment_amount', 0)):,.2f} | "
        f"Trials={inputs.get('trials')} | Jobs={inputs.get('jobs')} | Min Trades={inputs.get('min_trades')}"
    )


def result_inputs_differ_from_form(result_data: Dict[str, Any], form_inputs: Dict[str, Any]) -> bool:
    result_inputs = result_data.get("inputs") or {}
    if not result_inputs:
        return str(result_data.get("symbol", "")).upper() != str(form_inputs.get("ticker", "")).upper()
    keys = ["ticker", "start_date", "end_date", "investment_amount", "timeframe", "trials", "jobs", "min_trades"]
    return any(str(result_inputs.get(k)) != str(form_inputs.get(k)) for k in keys)


def risk_filter_for_signal(sig: Dict[str, Any], summary: Dict[str, Any]) -> Dict[str, str]:
    """Simple UI-level risk read, without changing the strategy engine."""
    status = str(sig.get("current_status") or "").upper()
    has_open = status == "OPEN" or bool(summary.get("has_open_trade"))

    if not has_open:
        return {
            "label": "NO_OPEN_TRADE",
            "level": "wait",
            "message": "No open trade. Wait for a new valid signal.",
        }

    pnl = sig.get("current_pnl_pct")
    risk = current_distance_to_stop_pct(sig)
    reward = current_distance_to_target_pct(sig)
    action = str(sig.get("action") or "").upper()

    def to_float(x):
        try:
            return float(x)
        except Exception:
            return None

    pnl_f = to_float(pnl)
    risk_f = to_float(risk)
    reward_f = to_float(reward)

    if "NEAR_STOP" in action or (risk_f is not None and abs(risk_f) <= 2.0):
        return {
            "label": "NEAR_STOP_WARNING",
            "level": "bad",
            "message": "The trade is close to the stop. Avoid adding risk and monitor the exit.",
        }

    if pnl_f is not None and pnl_f < -3:
        return {
            "label": "LOSING_TRADE_CAUTION",
            "level": "warn",
            "message": "The trade is currently losing. Follow the stop instead of adjusting randomly.",
        }

    if reward_f is not None and reward_f <= 3 and pnl_f is not None and pnl_f > 0:
        return {
            "label": "NEAR_TARGET",
            "level": "good",
            "message": "The trade is close to target. Follow the plan and trailing stop.",
        }

    if pnl_f is not None and pnl_f >= 5:
        return {
            "label": "HOLD_PROFIT",
            "level": "good",
            "message": "The trade is profitable. Keep following the plan.",
        }

    return {
        "label": "NORMAL_RISK",
        "level": "wait",
        "message": "Risk is normal based on the available data. Follow the plan.",
    }


def render_risk_filter(sig: Dict[str, Any], summary: Dict[str, Any]) -> None:
    risk = risk_filter_for_signal(sig, summary)
    css_class = {
        "good": "decision-good",
        "warn": "decision-warn",
        "bad": "decision-bad",
        "wait": "decision-wait",
    }.get(risk["level"], "decision-wait")
    st.markdown("### Risk Filter")
    st.markdown(f'<div class="{css_class}">{risk["label"]}</div>', unsafe_allow_html=True)
    st.caption(risk["message"])





def write_json_atomic(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    tmp.replace(path)


def load_current_job() -> Optional[Dict[str, Any]]:
    if not JOB_FILE.exists():
        return None
    data = safe_load_json(JOB_FILE)
    return data if isinstance(data, dict) and data else None


def clear_current_job() -> None:
    try:
        if JOB_FILE.exists():
            JOB_FILE.unlink()
    except Exception:
        pass


def pid_is_running(pid: Any) -> bool:
    try:
        pid_int = int(pid)
    except Exception:
        return False

    proc = st.session_state.get("current_job_process")
    if proc is not None:
        try:
            if int(getattr(proc, "pid", -1)) == pid_int:
                return proc.poll() is None
        except Exception:
            pass

    try:
        if os.name == "nt":
            result = subprocess.run(
                ["tasklist", "/FI", f"PID eq {pid_int}", "/FO", "CSV", "/NH"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="ignore",
                timeout=5,
            )
            return str(pid_int) in result.stdout
        os.kill(pid_int, 0)
        return True
    except Exception:
        return False


def read_job_log(max_lines: int = 160) -> str:
    if not JOB_LOG_FILE.exists():
        return ""
    try:
        lines = JOB_LOG_FILE.read_text(encoding="utf-8", errors="replace").splitlines()
        return "\n".join(lines[-max_lines:])
    except Exception:
        return ""


def trial_progress_from_log(log: str, trials: Optional[int]) -> float:
    if not trials:
        return 0.08
    found = [int(x) for x in re.findall(r"\[\s*(\d+)\]", log or "")]
    if not found:
        if "Loading data" in log:
            return 0.05
        if "Starting Optimization" in log:
            return 0.08
        return 0.02
    return min(0.95, max(0.02, max(found) / max(int(trials), 1)))


def write_background_runner() -> None:
    runner_code = r"""
from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path


def read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def write_json_atomic(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)


def newest(folder: Path, patterns: list[str], since_ts: float | None = None) -> str | None:
    files = []
    for pat in patterns:
        files.extend(folder.glob(pat))
    if since_ts is not None:
        files = [p for p in files if p.exists() and p.stat().st_mtime >= since_ts]
    if not files:
        return None
    return str(max(files, key=lambda p: p.stat().st_mtime))


def run_command(cmd: list[str], cwd: Path, log_file: Path, title: str) -> int:
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    with open(log_file, "a", encoding="utf-8", errors="replace") as log:
        log.write("\n\n" + "=" * 72 + "\n")
        log.write(title + "\n")
        log.write("COMMAND: " + " ".join(str(x) for x in cmd) + "\n")
        log.write("=" * 72 + "\n")
        log.flush()
        proc = subprocess.Popen(
            cmd,
            cwd=str(cwd),
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
        )
        rc = proc.wait()
        log.write(f"\n[{datetime.now().isoformat(timespec='seconds')}] EXIT CODE: {rc}\n")
        log.flush()
        return rc


def main() -> int:
    job_file = Path(sys.argv[1])
    job = read_json(job_file)
    if not job:
        return 2

    base_dir = Path(job["base_dir"])
    results_dir = Path(job["results_dir"])
    state_dir = Path(job["state_dir"])
    log_file = Path(job["log_file"])
    core_script = Path(job["core_script"])
    symbol = str(job["symbol"]).upper()
    started_ts = float(job.get("started_ts") or time.time())

    def update(status: str, **kwargs) -> None:
        latest = read_json(job_file) or job
        latest.update(kwargs)
        latest["status"] = status
        latest["updated_at"] = datetime.now().isoformat(timespec="seconds")
        write_json_atomic(job_file, latest)

    try:
        update("running", phase="optimize", pid=os.getpid())

        opt_cmd = [sys.executable, "-u", str(core_script)] + job["optimize_args"]
        rc = run_command(opt_cmd, base_dir, log_file, f"Optimize Log - {symbol}")
        if rc != 0:
            update("failed", phase="optimize", returncode=rc, error="Optimize failed. Open the log for details.")
            return rc

        opt_json = newest(results_dir, [f"{symbol}_best_strategy_*.json"], since_ts=started_ts)
        opt_xlsx = newest(results_dir, [f"{symbol}_best_strategy_*.xlsx"], since_ts=started_ts)
        if not opt_json:
            update("failed", phase="optimize", error="Optimize finished but no JSON was found in results/.")
            return 3

        update("running", phase="signal", opt_json=opt_json, opt_xlsx=opt_xlsx)

        signal_started_ts = time.time()
        signal_args = list(job["signal_args"]) + ["--params-file", opt_json]
        sig_cmd = [sys.executable, "-u", str(core_script)] + signal_args
        sig_rc = run_command(sig_cmd, base_dir, log_file, f"Signal Log - {symbol}")

        signal_json = None
        signal_xlsx = None
        signal_payload = {}
        sig = {}
        if sig_rc == 0:
            signal_json = newest(results_dir, [f"{symbol}_signal_*.json"], since_ts=signal_started_ts)
            signal_xlsx = newest(results_dir, [f"{symbol}_signal_*.xlsx"], since_ts=signal_started_ts)
            if signal_json:
                signal_payload = read_json(Path(signal_json))
                sig = signal_payload.get("current_signal", {}) or {}

        payload = read_json(Path(opt_json))
        summary = {}
        summary.update(payload.get("summary", {}) or {})
        summary.update(signal_payload.get("summary", {}) or {})

        analysis_data = {
            "symbol": symbol,
            "inputs": job.get("inputs", {}),
            "summary": summary,
            "current_signal": sig,
            "opt_json": opt_json,
            "opt_xlsx": opt_xlsx,
            "signal_json": signal_json,
            "signal_xlsx": signal_xlsx,
            "package": "",
            "saved_plan": "",
        }

        last_analysis_file = state_dir / "last_analysis.json"
        write_json_atomic(last_analysis_file, analysis_data)

        if sig_rc != 0:
            update(
                "done",
                phase="signal_failed",
                warning="Optimize completed, but Signal failed. The result was saved; check the log.",
                signal_returncode=sig_rc,
                analysis=analysis_data,
            )
        else:
            update("done", phase="complete", returncode=0, analysis=analysis_data)

        return 0

    except Exception as exc:
        update("failed", phase="exception", error=repr(exc))
        return 99


if __name__ == "__main__":
    raise SystemExit(main())
"""
    JOB_RUNNER_FILE.write_text(runner_code, encoding="utf-8")


def start_background_analysis_job(inputs: Dict[str, Any]) -> Dict[str, Any]:
    core = find_core_script()
    if core is None:
        raise RuntimeError("Core script not found. Expected wealth_engine.py or script.py")

    symbol = str(inputs["ticker"]).strip().upper()
    started_ts = time.time()
    job_id = datetime.now().strftime("%Y%m%d_%H%M%S") + f"_{symbol}"

    optimize_args = [
        symbol, "--mode", "optimize",
        "--start-date", str(inputs["start_date"]), "--end-date", str(inputs["end_date"]),
        "--investment-amount", str(float(inputs["investment_amount"])),
        "--timeframes", str(inputs["timeframe"]),
        "--trials", str(int(inputs["trials"])), "--jobs", str(int(inputs["jobs"])),
        "--min-trades", str(int(inputs["min_trades"])),
        "--adjust", "splits", "--output", str(RESULTS_DIR),
    ]

    signal_args = [
        symbol, "--mode", "signal",
        "--start-date", str(inputs["start_date"]), "--end-date", str(inputs["end_date"]),
        "--investment-amount", str(float(inputs["investment_amount"])),
        "--adjust", "splits", "--output", str(RESULTS_DIR),
    ]

    job = {
        "job_id": job_id,
        "symbol": symbol,
        "status": "starting",
        "phase": "starting",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "started_ts": started_ts,
        "base_dir": str(BASE_DIR),
        "results_dir": str(RESULTS_DIR),
        "state_dir": str(STATE_DIR),
        "log_file": str(JOB_LOG_FILE),
        "core_script": str(core),
        "inputs": inputs,
        "optimize_args": optimize_args,
        "signal_args": signal_args,
    }

    write_json_atomic(JOB_FILE, job)
    JOB_LOG_FILE.write_text(f"[{job['started_at']}] Starting background analysis for {symbol}\n", encoding="utf-8")
    write_background_runner()

    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"

    proc = subprocess.Popen(
        [sys.executable, str(JOB_RUNNER_FILE), str(JOB_FILE)],
        cwd=str(BASE_DIR),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )

    job["pid"] = proc.pid
    job["status"] = "running"
    write_json_atomic(JOB_FILE, job)
    st.session_state["current_job_process"] = proc
    return job


def cancel_background_job(job: Dict[str, Any]) -> None:
    pid = job.get("pid")
    proc = st.session_state.get("current_job_process")
    try:
        if proc is not None and proc.poll() is None:
            proc.terminate()
        elif pid:
            if os.name == "nt":
                subprocess.run(["taskkill", "/PID", str(int(pid)), "/T", "/F"], capture_output=True, text=True, timeout=10)
            else:
                os.kill(int(pid), 15)
    except Exception:
        pass
    job["status"] = "cancelled"
    job["updated_at"] = datetime.now().isoformat(timespec="seconds")
    write_json_atomic(JOB_FILE, job)


def consume_completed_job_if_needed(job: Dict[str, Any]) -> None:
    if not job or job.get("status") != "done":
        return
    job_id = job.get("job_id")
    if st.session_state.get("last_consumed_job_id") == job_id:
        return

    analysis = job.get("analysis") or safe_load_json(STATE_DIR / "last_analysis.json")
    if analysis:
        st.session_state["last_analysis"] = analysis
        persist_last_analysis(analysis)
        st.session_state["last_consumed_job_id"] = job_id


def render_background_job_status() -> Optional[Dict[str, Any]]:
    job = load_current_job()
    if not job:
        return None

    status = str(job.get("status", "")).lower()
    log = read_job_log()
    trials = None
    try:
        trials = int((job.get("inputs") or {}).get("trials") or 0)
    except Exception:
        trials = None

    if status == "running" and job.get("pid") and not pid_is_running(job.get("pid")):
        # Give the runner a chance to write final status; if output exists, do not force failure immediately.
        refreshed = load_current_job() or job
        if str(refreshed.get("status", "")).lower() == "running":
            refreshed["status"] = "failed"
            refreshed["error"] = "Background process stopped before writing a final result. Open the log for details."
            refreshed["updated_at"] = datetime.now().isoformat(timespec="seconds")
            write_json_atomic(JOB_FILE, refreshed)
            job = refreshed
            status = "failed"

    if status in ("running", "starting"):
        progress = trial_progress_from_log(log, trials)
        phase = job.get("phase", "running")
        symbol = job.get("symbol", "-")
        st.info(f"Running {symbol} • {phase}")
        st.progress(progress, text=f"{symbol} analysis... {progress * 100:.0f}%")
        cols = st.columns([1, 1, 3])
        with cols[0]:
            if st.button("🔄 Refresh Signal"):
                st.rerun()
        with cols[1]:
            if st.button("Cancel Job"):
                cancel_background_job(job)
                st.rerun()
        with st.expander(f"Live Log - {symbol}", expanded=False):
            st.code(log or "No log yet.")
        return job

    if status == "done":
        consume_completed_job_if_needed(job)
        st.markdown(f'<div class="clean-success">Analysis complete: {job.get("symbol", "-")}</div>', unsafe_allow_html=True)
        if job.get("warning"):
            st.warning(job.get("warning"))
        with st.expander("Completed Job Log", expanded=False):
            st.code(log or "No log.")
        return job

    if status in ("failed", "cancelled"):
        msg = job.get("error") or ("Analysis cancelled." if status == "cancelled" else "Analysis failed.")
        st.error(msg)
        with st.expander("Failed Job Log", expanded=True):
            st.code(log or "No log.")
        cols = st.columns([1, 4])
        with cols[0]:
            if st.button("Clear Job"):
                clear_current_job()
                st.session_state["current_job_process"] = None
                st.rerun()
        return job

    return job



def page_analyze_stock() -> None:
    st.markdown('<div class="wealth-title">Analyze Stock</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Run a new optimization. Save the result to My Stock or Watchlist.</div>', unsafe_allow_html=True)

    last_inputs = load_last_inputs()
    timeframe_options = ["2h", "4h", "1day"]
    default_tf = last_inputs.get("timeframe", "4h")
    default_tf_index = timeframe_options.index(default_tf) if default_tf in timeframe_options else 1

    default_end = _date_from_input(last_inputs.get("end_date"), date.today())
    default_start = _date_from_input(last_inputs.get("start_date"), default_end - timedelta(days=365))

    col1, col2, col3 = st.columns(3)
    with col1:
        ticker = st.text_input("Ticker", value=str(last_inputs.get("ticker", "IREN")), key="analyze_ticker").strip().upper()
        timeframe = st.selectbox("Timeframe", timeframe_options, index=default_tf_index, key="analyze_timeframe")
    with col2:
        start_date = st.date_input("Start Date", value=default_start, key="analyze_start_date")
        trials = st.number_input("Trials", min_value=1, max_value=100000, value=int(last_inputs.get("trials", 500)), step=50, key="analyze_trials")
    with col3:
        end_date = st.date_input("End Date", value=default_end, key="analyze_end_date")
        jobs = st.number_input("Jobs", min_value=1, max_value=64, value=int(last_inputs.get("jobs", 8)), step=1, key="analyze_jobs")

    col4, col5 = st.columns(2)
    with col4:
        investment_amount = st.number_input("Investment Amount $", min_value=1.0, max_value=1_000_000_000.0, value=float(last_inputs.get("investment_amount", 100.0)), step=100.0, key="analyze_investment_amount")
    with col5:
        min_trades = st.number_input("Min Trades", min_value=1, max_value=500, value=int(last_inputs.get("min_trades", 12)), step=1, key="analyze_min_trades")

    if start_date > end_date:
        st.error("Start Date must be before or equal to End Date.")
        return

    form_inputs = current_form_inputs()

    job = render_background_job_status()
    job_status = str((job or {}).get("status", "")).lower()

    last = load_last_analysis()
    if last and result_inputs_differ_from_form(last, form_inputs):
        old_inputs = last.get("inputs") or {"ticker": last.get("symbol", "-")}
        st.warning(
            "The result below belongs to: "
            f"{inputs_for_display(old_inputs)}. "
            f"Current form: {inputs_for_display(form_inputs)}. "
            "Click Analyze Stock to run a new result for the current form."
        )

    analyze_disabled = job_status in ("running", "starting")
    if st.button("Analyze Stock", type="primary", use_container_width=True, disabled=analyze_disabled):
        if not ticker:
            st.warning("Enter a ticker first.")
            return

        analysis_inputs = current_form_inputs()
        save_last_inputs(analysis_inputs)

        try:
            start_background_analysis_job(analysis_inputs)
            st.success("Analysis started.")
            st.rerun()
        except Exception as exc:
            st.error(f"Could not start analysis: {exc}")
            return

    if analyze_disabled:
        st.caption("A job is running. Wait until it finishes or cancel it first.")

    last = load_last_analysis()

    if last:
        symbol, summary, sig = last["symbol"], last.get("summary", {}), last.get("current_signal", {})
        result_inputs = last.get("inputs") or {"ticker": symbol}
        st.markdown(
            f'<div class="clean-success">Analysis complete: {symbol} • TF={result_inputs.get("timeframe", "-")} • Save it below.</div>',
            unsafe_allow_html=True,
        )

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Investment Amount", "$" + fmt_num(summary.get("investment_amount", result_inputs.get("investment_amount")), 2))
        c2.metric("Final Capital", "$" + fmt_num(summary.get("final_capital"), 2))
        c3.metric("Compound Profit %", fmt_signed_pct(summary_compound_pct(summary)))
        c4.metric("Strategy P/L", fmt_signed_pct(summary_strategy_pl(summary)))

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Trades", get_first(summary, ["num_trades", "trades"], "-"))
        c6.metric("Win Rate", fmt_abs_pct(get_first(summary, ["win_rate"], None)))
        c7.metric("Max Drawdown", fmt_abs_pct(get_first(summary, ["max_drawdown_pct"], None)))
        c8.metric("Open Trade", "YES" if (sig.get("current_status") == "OPEN" or summary.get("has_open_trade")) else "NO")

        if sig:
            st.subheader("Current Signal")
            decision = sig.get("action", "-")
            st.markdown(f'<div class="{decision_class(decision)}">Smart Decision: {decision}</div>', unsafe_allow_html=True)
            st.caption(signal_short_reason(sig))

            oc1, oc2, oc3, oc4 = st.columns(4)
            oc1.metric("Entry", fmt_num(sig.get("last_entry_price")))
            oc2.metric("Current", fmt_num(sig.get("current_price")))
            oc3.metric("Stop", fmt_num(sig.get("current_stop")))
            oc4.metric("Target", fmt_num(sig.get("current_target")))

            oc5, oc6, oc7 = st.columns(3)
            display_risk, display_reward = entry_based_trade_percents(sig)
            oc5.metric("Current PnL", fmt_signed_pct(sig.get("current_pnl_pct")))
            oc6.metric("Risk to Stop", fmt_signed_pct(display_risk))
            oc7.metric("Reward to Target", fmt_signed_pct(display_reward))

        render_risk_filter(sig, summary)

        st.subheader("Save Result")
        a1, a2, a3 = st.columns([1.3, 1.3, 1])

        with a1:
            if st.button(f"Save to My Stock", use_container_width=True):
                save_last_to_category(symbol, last, "my_stock")

        with a2:
            if st.button(f"Save to Watchlist", use_container_width=True):
                save_last_to_category(symbol, last, "watchlist")

        with a3:
            if last.get("opt_xlsx"):
                download_button_for_file(Path(last["opt_xlsx"]), "Download Excel", key_prefix="analyze_excel")

        package_value = st.session_state.get("last_package") or last.get("package") or ""
        package = Path(package_value) if package_value else None
        if package and package.exists() and package.is_file():
            download_button_for_file(package, f"Download {package.name}", key_prefix="analyze_package")


def page_dashboard(title: str, symbols: List[str], group_key: str) -> None:
    st.markdown(f'<div class="wealth-title">{title}</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Portfolio summary for saved plans in this section.</div>', unsafe_allow_html=True)

    param_files = plan_files_for_symbols(symbols)
    if not param_files:
        st.info("No saved plans in this section yet.")
        return

    raw_df = pd.DataFrame([dashboard_row(p) for p in param_files])

    rename_map = {"Has Open Trade": "Open Trade"}
    raw_df = raw_df.rename(columns=rename_map)
    dashboard_cols = [
        "Symbol",
        "Period",
        "Timeframe",
        "Smart Decision",
        "Open Trade",
        "Investment Amount",
        "Final Capital",
        "Compound Profit %",
        "Strategy P/L",
        "Win Rate",
        "Max Drawdown",
        "Trades",
    ]
    df = raw_df[[c for c in dashboard_cols if c in raw_df.columns]]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Symbols", len(param_files))
    c2.metric("Open Trades", int((df["Open Trade"] == "YES").sum()) if "Open Trade" in df.columns else 0)
    c3.metric("Near Stop", int(df["Smart Decision"].astype(str).str.contains("NEAR_STOP", na=False).sum()) if "Smart Decision" in df.columns else 0)
    c4.metric("Waiting", int(df["Smart Decision"].astype(str).str.contains("WAIT", na=False).sum()) if "Smart Decision" in df.columns else 0)

    col1, col2, col3, col4 = st.columns([1.15, 2.35, 0.42, 1.05])
    with col1:
        years = st.number_input("Refresh years", min_value=1, max_value=20, value=int(st.session_state.get(f"{group_key}_refresh_years", 1)), step=1, key=f"{group_key}_years")
        st.session_state[f"{group_key}_refresh_years"] = int(years)
    with col2:
        st.markdown('<div class="control-row-spacer"></div>', unsafe_allow_html=True)
        if st.button("Refresh Signals", type="primary", use_container_width=True, key=f"refresh_{group_key}"):
            for p in param_files:
                run_signal_for_plan(p, years=int(years))
            st.success("Signals refreshed.")
            st.rerun()
    with col3:
        st.markdown('<div class="control-row-spacer"></div>', unsafe_allow_html=True)
        download_dataframe_csv(df, f"{group_key}_dashboard.csv", "↓ CSV", f"download_{group_key}_dashboard_csv")

    st.markdown("### Summary")
    st.dataframe(style_dashboard_df(df), use_container_width=True, hide_index=True)


def page_trades_overview(title: str, symbols: List[str], group_key: str) -> None:
    st.markdown(f'<div class="wealth-title">{title}</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Current trade status for this section.</div>', unsafe_allow_html=True)

    if not symbols:
        st.info("No saved plans in this section yet.")
        return

    rows = []
    missing_signal = []

    for symbol in symbols:
        plan_file = PLANS_DIR / f"{symbol}_plan.json"
        plan_payload = safe_load_json(plan_file) if plan_file.exists() else {}
        plan_summary = plan_payload.get("summary", {}) if plan_payload else {}

        signal_payload = extract_signal_payload(symbol)
        if not signal_payload:
            missing_signal.append(symbol)
            sig = {}
            sig_summary = {}
        else:
            sig = signal_payload.get("current_signal", {})
            sig_summary = signal_payload.get("summary", {})

        summary = {**plan_summary, **sig_summary}
        status = sig.get("current_status") or ("OPEN" if summary.get("has_open_trade") else "CLOSED")
        has_open = status == "OPEN"
        decision = sig.get("action") or ("HOLD" if has_open else "WAIT")

        display_risk, display_reward = entry_based_trade_percents(sig) if has_open else (None, None)

        rows.append({
            "Symbol": symbol,
            "Open Trade": "YES" if has_open else "NO",
            "Smart Decision": decision,
            "Entry Price": fmt_num(sig.get("last_entry_price")) if has_open else "-",
            "Current Price": fmt_num(sig.get("current_price")),
            "Current PnL %": fmt_signed_pct(sig.get("current_pnl_pct")) if has_open else "-",
            "Stop": fmt_num(sig.get("current_stop")) if has_open else "-",
            "Target": fmt_num(sig.get("current_target")) if has_open else "-",
            "Risk to Stop %": fmt_signed_pct(display_risk) if has_open else "-",
            "Reward to Target %": fmt_signed_pct(display_reward) if has_open else "-",
            "Timeframe": timeframe_from_payload(signal_payload or plan_payload or {}),
            "Period": summary.get("analysis_period") or f"{summary.get('analysis_start', '')} -> {summary.get('analysis_end', '')}",
            "Investment Amount": "$" + fmt_num(summary.get("investment_amount", summary.get("initial_capital")), 2),
            "Final Capital": "$" + fmt_num(summary.get("final_capital"), 2),
            "Compound Profit %": fmt_signed_pct(summary_compound_pct(summary)),
            "Strategy P/L": fmt_signed_pct(summary_strategy_pl(summary)),
            "Win Rate": fmt_abs_pct(get_first(summary, ["win_rate"], None)),
            "Trades": get_first(summary, ["num_trades", "trades"], "-"),
        })

    if missing_signal:
        st.warning("No Signal file yet for: " + ", ".join(missing_signal) + ". Refresh this section or use Signals page.")

    df = pd.DataFrame(rows)

    total_symbols = len(df)
    open_count = int((df["Open Trade"] == "YES").sum()) if "Open Trade" in df.columns else 0
    waiting_count = int((df["Open Trade"] == "NO").sum()) if "Open Trade" in df.columns else 0
    near_stop_count = int(df["Smart Decision"].astype(str).str.contains("NEAR_STOP", na=False).sum()) if "Smart Decision" in df.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Symbols", total_symbols)
    c2.metric("Open Trades", open_count)
    c3.metric("Waiting", waiting_count)
    c4.metric("Near Stop", near_stop_count)

    title_col, csv_col, _ = st.columns([2.2, 0.42, 2.1])
    with title_col:
        st.subheader("Portfolio / Current Trade Status")
    with csv_col:
        st.markdown('<div class="control-row-spacer"></div>', unsafe_allow_html=True)
        download_dataframe_csv(df, f"{group_key}_trades.csv", "↓ CSV", f"download_{group_key}_trades_csv")

    st.dataframe(style_dashboard_df(df), use_container_width=True, hide_index=True)

    st.caption("Open the Signals page for full symbol details.")

def _compact_trade_table(df: pd.DataFrame, limit: int = 4) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    sort_col = "entry_time" if "entry_time" in out.columns else ("Entry Time" if "Entry Time" in out.columns else None)
    if sort_col:
        out[sort_col] = pd.to_datetime(out[sort_col], errors="coerce")
        out = out.sort_values(sort_col)
    out = out.tail(limit)
    rename_map = {
        "trade_no": "#",
        "timeframe": "TF",
        "status": "Status",
        "entry_time": "Entry Time",
        "entry_price": "Entry",
        "exit_time": "Exit Time",
        "exit_price": "Exit",
        "reason": "Reason",
        "current_price": "Current",
        "current_stop": "Stop",
        "current_target": "Target",
        "pnl_pct": "PnL %",
        "trade_return_pct": "Return %",
    }
    out = out.rename(columns=rename_map)
    preferred = ["#", "TF", "Status", "Entry Time", "Entry", "Exit Time", "Exit", "Reason", "Current", "Stop", "Target", "PnL %", "Return %"]
    keep = [c for c in preferred if c in out.columns]
    if keep:
        out = out[keep]
    for col in ["Entry Time", "Exit Time"]:
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
    for col in ["Entry", "Exit", "Current", "Stop", "Target", "PnL %", "Return %"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").round(2)
    return out


def page_signal_detail(title: str, symbols: List[str], group_key: str) -> None:
    st.markdown(f'<div class="wealth-title">{title}</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Current decision and latest 4 trades for this section.</div>', unsafe_allow_html=True)
    if not symbols:
        st.info("No saved plans in this section yet.")
        return

    symbol = st.selectbox("Select Symbol", symbols, key=f"signal_symbol_{group_key}")
    plan_file = active_param_file(symbol)
    if not plan_file:
        st.warning("No usable plan found for this symbol.")
        return

    col1, col2, col3, col4 = st.columns([0.95, 1.95, 1.95, 0.8])
    with col1:
        years = st.number_input("Years", min_value=1, max_value=20, value=1, step=1, key=f"detail_years_{group_key}")
    with col2:
        st.markdown('<div class="control-row-spacer"></div>', unsafe_allow_html=True)
        if st.button("Refresh Signal", type="primary", use_container_width=True, key=f"refresh_signal_{group_key}"):
            run_signal_for_plan(plan_file, years=int(years))
            st.rerun()
    with col3:
        st.markdown('<div class="control-row-spacer"></div>', unsafe_allow_html=True)
        if st.button("Run Replay", use_container_width=True, key=f"run_replay_{group_key}"):
            run_replay_for_plan(plan_file, years=int(years))
            st.rerun()

    signal_payload = extract_signal_payload(symbol)
    replay_payload = extract_replay_payload(symbol)

    if not signal_payload:
        st.warning("No Signal saved yet. Click Refresh Signal.")
        return

    signal_summary = signal_payload.get("summary", {})
    sig = signal_payload.get("current_signal", {})
    replay_summary = replay_payload.get("summary", {}) if replay_payload else {}
    replay_events = replay_payload.get("recent_events", []) if replay_payload else []
    signal_events = signal_payload.get("recent_events", [])
    events = replay_events or signal_events
    events_source = "REPLAY" if replay_events else "SIGNAL"

    decision = sig.get("action", "-")
    st.markdown(f"### {symbol} / {signal_summary.get('timeframe') or timeframe_from_payload(signal_payload)}")
    st.markdown(f'<div class="{decision_class(decision)}">Decision: {decision}</div>', unsafe_allow_html=True)
    reason_text = signal_short_reason(sig)
    if reason_text:
        st.caption(reason_text)
    source_caption = f"Signal: SIGNAL · Trades: {events_source}"
    if replay_payload:
        source_caption += f" · Replay last bar: {replay_summary.get('last_bar_time', '-')}"
    st.caption(source_caption)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Entry", fmt_num(sig.get("last_entry_price")))
    c2.metric("Current", fmt_num(sig.get("current_price")))
    c3.metric("Stop", fmt_num(sig.get("current_stop")))
    c4.metric("Target", fmt_num(sig.get("current_target")))

    c5, c6, c7 = st.columns(3)
    display_risk, display_reward = entry_based_trade_percents(sig)
    c5.metric("Current PnL", fmt_signed_pct(sig.get("current_pnl_pct")))
    c6.metric("Risk to Stop", fmt_signed_pct(display_risk))
    c7.metric("Reward to Target", fmt_signed_pct(display_reward))

    st.subheader("Latest 4 Trades")
    trades_df, trades_sheet_source = load_report_trades_df(symbol)
    compact = _compact_trade_table(trades_df, 4) if not trades_df.empty else pd.DataFrame()
    if compact.empty and events:
        compact = _compact_trade_table(pd.DataFrame(events), 4)
        trades_sheet_source = events_source
    if compact.empty:
        st.info("No trade history found yet. Run Replay to rebuild it.")
    else:
        st.caption(f"Showing latest rows from {trades_sheet_source}.")
        st.dataframe(compact, use_container_width=True, hide_index=True)

    replay_files = latest_replay_files(symbol) if replay_payload else {"xlsx": None, "json": None}
    package = latest_package(symbol)
    d1, d2, d3 = st.columns(3)
    with d1:
        if replay_files.get("xlsx"):
            download_button_for_file(replay_files["xlsx"], f"↓ {symbol}_REPLAY.xlsx", key_prefix="signal_replay_xlsx")
    with d2:
        if replay_files.get("json"):
            download_button_for_file(replay_files["json"], f"↓ {symbol}_REPLAY.json", key_prefix="signal_replay_json")
    with d3:
        if package and package.exists() and package.is_file():
            download_button_for_file(package, f"↓ {package.name}", key_prefix="signal_package")
    if not replay_payload:
        st.caption("Run Replay to show full trade history. Optimize stays unchanged.")

def package_status_for_symbol(symbol: str) -> Dict[str, Any]:
    symbol = symbol.upper()
    return {
        "Symbol": symbol,
        "Package": latest_package(symbol).name if latest_package(symbol) else "-",
        "Plan": "YES" if (PLANS_DIR / f"{symbol}_plan.json").exists() else "Missing",
        "Optimize": "YES" if (latest_report(symbol, "OPTIMIZE", "xlsx") or latest_report(symbol, "OPTIMIZE", "json")) else "Missing",
        "Signal": "YES" if (latest_report(symbol, "SIGNAL", "xlsx") or latest_report(symbol, "SIGNAL", "json")) else "Missing",
        "Replay": "YES" if (latest_report(symbol, "REPLAY", "xlsx") or latest_report(symbol, "REPLAY", "json")) else "Missing",
    }


def page_files_reports() -> None:
    st.markdown('<div class="wealth-title">Files / Reports</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Download and manage saved plan files.</div>', unsafe_allow_html=True)

    symbols = active_plan_symbols()
    if not symbols:
        st.info("No active plans yet. Save a plan from Analyze Stock first.")
        return

    symbol = st.selectbox("Select Symbol", symbols)

    if st.button(f"Refresh {symbol} Package", type="primary"):
        package_path, availability = create_symbol_package(symbol, "OPTIMIZE", force_new=False)
        st.success(f"Package refreshed: {package_path.name}")

    status_row = package_status_for_symbol(symbol)
    st.dataframe(pd.DataFrame([status_row]), use_container_width=True, hide_index=True)

    if status_row.get("Plan") == "Missing":
        st.warning(f"Plan is missing. Save this symbol from Analyze Stock to include {symbol}_PLAN.json in the package.")

    package = latest_package(symbol)
    if package and package.exists() and package.is_file():
        st.subheader("Download Package")
        pkg_col, section_col = st.columns([1.25, 1.0])
        with pkg_col:
            st.caption(f"Package: `{package.name}`")
            download_button_for_file(package, f"↓ Download {package.name}", key_prefix="files_package")
        with section_col:
            current_section = symbol_section(symbol)
            target_section = "Watchlist" if current_section == "My Stock" else "My Stock"
            st.caption("Section")
            st.markdown(
                f"""
                <div class="section-card">
                    <div><span class="muted-label">Symbol</span><strong>{symbol}</strong></div>
                    <div><span class="muted-label">Current</span><strong>{current_section}</strong></div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button(f"Move to {target_section}", use_container_width=True, key=f"move_section_{symbol}_{target_section}"):
                moved_to = move_symbol_to_section(symbol, target_section)
                st.success(f"{symbol} moved to {moved_to}.")
                st.rerun()
    else:
        st.info("No package yet. Click Refresh Package.")
        current_section = symbol_section(symbol)
        target_section = "Watchlist" if current_section == "My Stock" else "My Stock"
        st.caption(f"{symbol} is currently in {current_section}.")
        if st.button(f"Move to {target_section}", use_container_width=True, key=f"move_section_{symbol}_{target_section}_nopkg"):
            moved_to = move_symbol_to_section(symbol, target_section)
            st.success(f"{symbol} moved to {moved_to}.")
            st.rerun()

    with st.expander("Show all official files", expanded=False):
        raw_files = official_symbol_files(symbol)
        if raw_files:
            rows = [
                {
                    "File": p.name,
                    "Folder": p.parent.name,
                    "Modified": datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                    "Size KB": round(p.stat().st_size / 1024, 1),
                }
                for p in raw_files
            ]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.write("No official files.")

    st.markdown("---")
    st.markdown("### Danger Zone")
    st.markdown('<div class="danger-zone">', unsafe_allow_html=True)
    st.warning("Remove Active Plan moves files to archive/. Nothing is permanently deleted.")
    confirm = st.text_input(f"Type {symbol} to confirm removal", key=f"remove_confirm_{symbol}")
    if st.button(f"Archive / Remove {symbol} Active Plan", type="secondary", disabled=(confirm.strip().upper() != symbol)):
        archive_path = archive_active_symbol(symbol)
        st.success(f"{symbol} moved to archive: {archive_path.name}")
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


def page_auto_refresh() -> None:
    st.markdown('<div class="wealth-title">Auto Refresh</div>', unsafe_allow_html=True)
    st.markdown('<div class="wealth-subtitle">Refresh saved plan signals. Use Safe Mode for free Twelve Data limits.</div>', unsafe_allow_html=True)

    plan_files = active_param_files()
    active_symbols = sorted({symbol_from_json(p) for p in plan_files})
    active_count = len(active_symbols)

    c1, c2, c3 = st.columns(3)
    c1.metric("Saved Plans", active_count)
    c2.metric("My Stock", len([s for s in active_symbols if s in load_my_stocks()]))
    c3.metric("Watchlist", len([s for s in active_symbols if s in load_watchlist()]))

    if not active_symbols:
        st.info("No saved plans yet.")
        return

    st.subheader("Covered Symbols")
    st.dataframe(pd.DataFrame({"Symbol": active_symbols}), use_container_width=True, hide_index=True)

    st.markdown("### Refresh Mode")
    mode = st.radio(
        "Mode",
        ["Free API Safe Refresh", "Fast Refresh"],
        index=0,
        horizontal=True,
        help="Safe mode refreshes one symbol at a time and waits between symbols.",
    )

    years = st.number_input("Signal years", min_value=1, max_value=20, value=int(st.session_state.get("auto_refresh_years", 3)), step=1, key="auto_years")
    st.session_state["auto_refresh_years"] = int(years)

    if mode == "Free API Safe Refresh":
        s1, s2, s3, s4 = st.columns(4)
        api_limit = s1.number_input("Credits / minute", min_value=1, max_value=1000, value=int(st.session_state.get("safe_api_limit", 4)), step=1, key="safe_api_limit")
        request_pause = s2.number_input("Pause inside symbol", min_value=0.0, max_value=30.0, value=float(st.session_state.get("safe_request_pause", 4.0)), step=0.5, key="safe_request_pause")
        fallback_wait = s3.number_input("Wait after symbol", min_value=0, max_value=600, value=int(st.session_state.get("safe_fallback_wait", 20)), step=5, key="safe_fallback_wait")
        buffer_seconds = s4.number_input("Buffer seconds", min_value=0, max_value=120, value=int(st.session_state.get("safe_buffer_seconds", 4)), step=5, key="safe_buffer_seconds")

        st.caption("Safe mode slows down Twelve Data requests and waits between symbols to avoid the free-plan credit limit.")

        if st.button("Start Safe Refresh Queue", type="primary", use_container_width=True):
            total = len(plan_files)
            results = []
            queue_progress = st.progress(0, text=f"Starting queue: 0/{total}")
            for idx, plan_file in enumerate(plan_files, start=1):
                symbol = symbol_from_json(plan_file)
                queue_progress.progress((idx - 1) / max(total, 1), text=f"Refreshing {idx}/{total}: {symbol}")
                outputs = run_signal_for_plan(plan_file, years=int(years), pause_sec=float(request_pause))
                ok = bool(outputs.get("json"))
                results.append({"Symbol": symbol, "Status": "Updated" if ok else "Failed"})

                if idx < total:
                    wait_for = wait_seconds_from_api_log(
                        st.session_state.get("last_log", ""),
                        api_limit_per_minute=int(api_limit),
                        buffer_seconds=int(buffer_seconds),
                        fallback_seconds=int(fallback_wait),
                    )
                    countdown_wait(wait_for, label=f"API cooldown after {symbol}")

            queue_progress.progress(1.0, text="Safe refresh queue complete")
            st.success("Safe refresh complete.")
            st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)
            st.rerun()

    else:
        st.warning("Fast Refresh can hit the free API limit if you refresh many symbols.")
        if st.button("Refresh All Now", type="primary", use_container_width=True):
            for p in plan_files:
                run_signal_for_plan(p, years=int(years))
            st.success("Signals refreshed.")
            st.rerun()

    st.markdown("### Scheduled Auto Refresh")
    enabled = st.toggle("Enable scheduled refresh", value=bool(st.session_state.get("auto_refresh_enabled", False)), key="auto_enabled")
    interval = st.number_input("Refresh every X minutes", min_value=5, max_value=240, value=int(st.session_state.get("auto_refresh_interval", 30)), step=5, key="auto_interval")
    st.session_state["auto_refresh_enabled"] = bool(enabled)
    st.session_state["auto_refresh_interval"] = int(interval)

    if enabled:
        st.info(f"Scheduled refresh is ON every {int(interval)} minutes. For free API limits, use manual Safe Refresh Queue above.")
        run_auto_refresh_if_due(int(interval))
    else:
        st.info("Scheduled refresh is OFF.")

def _sidebar_nav_button(label: str, current_page: str) -> None:
    active = current_page == label
    if st.sidebar.button(label, key=f"nav_{label}", use_container_width=True, type="primary" if active else "secondary"):
        st.session_state["nav_page"] = label
        st.rerun()


def main() -> None:
    st.sidebar.markdown(f"## {APP_TITLE}")
    core = find_core_script()
    if core:
        st.sidebar.markdown(
            '<div class="sidebar-summary"><strong>Plan, track, and refresh stock strategies.</strong><br>'
            'Optimize entries, monitor current signals, and export reports from one place.</div>',
            unsafe_allow_html=True,
        )
    else:
        st.sidebar.error("Core script not found")
        st.sidebar.caption("Expected: wealth_engine.py")

    my_symbols = classified_symbols("my_stock")
    watch_symbols = classified_symbols("watchlist")

    valid_pages = [
        "Analyze Stock",
        "My Stock Dashboard",
        "Trades Of My Stock",
        "My Stock Signals",
        "Watchlist Dashboard",
        "Trades Of Watchlist",
        "Watchlist Signals",
        "Files / Reports",
        "Auto Refresh",
    ]
    page = st.session_state.get("nav_page", "Analyze Stock")
    if page not in valid_pages:
        page = "Analyze Stock"
        st.session_state["nav_page"] = page

    st.sidebar.markdown('<div class="nav-section">Workspace</div>', unsafe_allow_html=True)
    _sidebar_nav_button("Analyze Stock", page)

    st.sidebar.markdown('<div class="nav-section">My Stock</div>', unsafe_allow_html=True)
    _sidebar_nav_button("My Stock Dashboard", page)
    _sidebar_nav_button("Trades Of My Stock", page)
    _sidebar_nav_button("My Stock Signals", page)

    st.sidebar.markdown('<div class="nav-section">Watchlist</div>', unsafe_allow_html=True)
    _sidebar_nav_button("Watchlist Dashboard", page)
    _sidebar_nav_button("Trades Of Watchlist", page)
    _sidebar_nav_button("Watchlist Signals", page)

    st.sidebar.markdown('<div class="nav-section">Tools</div>', unsafe_allow_html=True)
    _sidebar_nav_button("Files / Reports", page)
    _sidebar_nav_button("Auto Refresh", page)

    st.sidebar.markdown("---")
    st.sidebar.markdown('<div class="nav-section">Portfolio</div>', unsafe_allow_html=True)
    st.sidebar.caption(f"My Stock: {len(my_symbols)}")
    st.sidebar.caption(f"Watchlist: {len(watch_symbols)}")
    st.sidebar.caption(f"Saved Plans: {len(active_plan_symbols())}")

    if page == "Analyze Stock":
        page_analyze_stock()
    elif page == "My Stock Dashboard":
        page_dashboard("My Stock Dashboard", my_symbols, "my_stock")
    elif page == "Trades Of My Stock":
        page_trades_overview("Trades Of My Stock", my_symbols, "my_stock")
    elif page == "My Stock Signals":
        page_signal_detail("My Stock Signals", my_symbols, "my_stock")
    elif page == "Watchlist Dashboard":
        page_dashboard("Watchlist Dashboard", watch_symbols, "watchlist")
    elif page == "Trades Of Watchlist":
        page_trades_overview("Trades Of Watchlist", watch_symbols, "watchlist")
    elif page == "Watchlist Signals":
        page_signal_detail("Watchlist Signals", watch_symbols, "watchlist")
    elif page == "Files / Reports":
        page_files_reports()
    elif page == "Auto Refresh":
        page_auto_refresh()


if __name__ == "__main__":
    main()