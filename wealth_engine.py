
#!/usr/bin/env python3
import argparse
import copy
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, TypedDict

import numpy as np
import optuna
import pandas as pd
import requests
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from optuna.samplers import TPESampler

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

API_BASE = "https://api.twelvedata.com"
DEFAULT_TWELVEDATA_API_KEY = "ec53cce5f4644d079413f09a5493f31b"

SUPPORTED_INTERVALS = ("2h", "4h", "1day")
INTRADAY_INTERVALS = {"2h", "4h"}

_ANNUALIZATION_DAYS = 252.0
_OPEN_TRADE_BONUS_CAP = 15.0
_OPEN_TRADE_BONUS_WEIGHT = 0.15
_PROFIT_FACTOR_FLOOR = 1e-9
_MIN_ATR_VALUE = 1e-6
_PROFIT_FACTOR_DISPLAY_CAP = 999.0
DEFAULT_INVESTMENT_AMOUNT = 100.0

ATR_MODE = "wilder"
RSI_FLAT_MODE = "neutral_50"

DEFAULT_PARAMS = {
    "initial_capital": 10000.0,
    "position_pct": 20.0,
    "commission_pct": 0.1,

    "use_htf": True,
    "htf_ema_len": 200,
    "ema_pull_len": 20,
    "ema_mid_len": 50,
    "ema_slow_len": 200,

    "entry_mode": "Both",
    "pullback_window": 12,
    "breakout_lookback": 20,
    "cooldown_bars": 3,

    "use_rsi": True,
    "rsi_len": 14,
    "rsi_min": 45,
    "rsi_max": 75,

    "use_stretch_filter": True,
    "max_stretch_atr": 1.5,

    "atr_len": 14,
    "sl_mult": 2.0,
    "tp_mult": 8.0,

    "use_be": True,
    "be_r": 1.0,

    "use_chand": True,
    "chand_mult": 4.0,

    "use_partial_tp": True,
    "partial_tp_r": 2.0,
    "partial_tp_pct": 50,

    "use_time_stop": True,
    "max_bars_in_trade": 30,
}

optuna.logging.set_verbosity(optuna.logging.WARNING)


class BacktestResult(TypedDict):
    net_profit_pct: float
    net_profit_value: float
    num_trades: int
    num_winning: int
    win_rate: float
    max_drawdown_pct: float
    profit_factor: float
    sharpe: float
    trades: list
    open_trade: Optional[dict]
    equity_curve: list
    last_bar_time: object
    last_close: float
    final_equity: float


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def parse_timeframes(value: str) -> List[str]:
    items = [x.strip() for x in value.split(",") if x.strip()]
    if not items:
        raise ValueError("No timeframes provided.")

    bad = [x for x in items if x not in SUPPORTED_INTERVALS]
    if bad:
        raise ValueError(
            f"Unsupported timeframes: {bad}. Allowed: {list(SUPPORTED_INTERVALS)}"
        )

    seen = set()
    out: List[str] = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _is_date_only(value: Optional[str]) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}|\d{4}/\d{2}/\d{2}", text))


def resolve_date_range(
    years: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Tuple[pd.Timestamp, pd.Timestamp]:
    """Resolve the requested analysis window.

    Fixes:
      - Date-only end dates are inclusive until 23:59:59.
      - Missing end_date means latest available / current timestamp.
      - Date-only start dates begin at 00:00:00.
    """
    if end_date:
        end_dt = pd.to_datetime(end_date, errors="raise").tz_localize(None)
        if _is_date_only(end_date):
            end_dt = end_dt.normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    else:
        end_dt = pd.Timestamp.now().tz_localize(None)

    if start_date:
        start_dt = pd.to_datetime(start_date, errors="raise").tz_localize(None)
        if _is_date_only(start_date):
            start_dt = start_dt.normalize()
    else:
        years = int(years or 1)
        start_dt = end_dt - relativedelta(years=years)

    if start_dt >= end_dt:
        raise ValueError("start_date must be before end_date.")

    return start_dt, end_dt

def analysis_period_label(years: Optional[int], start_date: Optional[str], end_date: Optional[str]) -> str:
    start_dt, end_dt = resolve_date_range(years=years, start_date=start_date, end_date=end_date)
    return f"{start_dt.date()} -> {end_dt.date()}"


def safe_float(value, default=np.nan) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


# ─────────────────────────────────────────────
# INDICATORS
# ─────────────────────────────────────────────

def ema(series: pd.Series, length: int) -> pd.Series:
    return series.ewm(span=length, adjust=False).mean()


def atr(df: pd.DataFrame, length: int) -> pd.Series:
    # FIX I2: Wilder ATR / RMA. This is the common TradingView-style ATR.
    high, low, close = df["high"], df["low"], df["close"]
    tr = pd.concat(
        [
            high - low,
            (high - close.shift(1)).abs(),
            (low - close.shift(1)).abs(),
        ],
        axis=1,
    ).max(axis=1)
    return tr.ewm(alpha=1.0 / length, min_periods=length, adjust=False).mean()


def rsi(series: pd.Series, length: int) -> pd.Series:
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = (-delta.clip(upper=0))
    avg_gain = gain.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    out = 100 - (100 / (1 + rs))
    # FIX: flat / undefined RSI starts neutral, not overbought.
    return out.fillna(50.0)


def annualization_factor(index: pd.Index) -> float:
    if len(index) < 20:
        return _ANNUALIZATION_DAYS

    normalized = pd.Series(index.normalize(), index=index)
    bars_per_day = normalized.groupby(normalized).size()
    median_bpd = float(bars_per_day.median()) if len(bars_per_day) else 1.0
    return max(_ANNUALIZATION_DAYS, _ANNUALIZATION_DAYS * median_bpd)


def align_htf_filter(df_htf: pd.DataFrame, df_ltf: pd.DataFrame, ema_len: int):
    # FIX: use completed HTF bars only. Prevents intraday bars from seeing
    # the same day's daily close/EMA before the daily candle completes.
    htf_close = df_htf["close"]
    htf_ema = ema(htf_close, ema_len)

    htf_close_safe = htf_close.shift(1)
    htf_ema_safe = htf_ema.shift(1)

    aligned_close = htf_close_safe.reindex(df_ltf.index, method="ffill")
    aligned_ema = htf_ema_safe.reindex(df_ltf.index, method="ffill")
    return aligned_close.values, aligned_ema.values


# ─────────────────────────────────────────────
# DATA FETCHING
# ─────────────────────────────────────────────

@dataclass
class DownloadConfig:
    api_key: str
    symbol: str
    years: int = 1
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    adjust: str = "splits"
    exchange: Optional[str] = None
    timezone: str = "Exchange"
    pause_sec: float = 0.35
    timeout_sec: int = 30
    retries: int = 3


class TwelveDataClient:
    def __init__(self, config: DownloadConfig):
        self.config = config
        self.session = requests.Session()
        self.session.headers.update({"Authorization": f"apikey {config.api_key}"})

    def _request(self, endpoint: str, params: Dict[str, str]) -> Dict:
        # FIX I8: retry with exponential backoff.
        url = f"{API_BASE}/{endpoint.lstrip('/')}"
        last_error = None

        for attempt in range(self.config.retries):
            try:
                response = self.session.get(
                    url,
                    params=params,
                    timeout=self.config.timeout_sec,
                )

                if response.status_code == 429:
                    raise RuntimeError("Rate limit / HTTP 429")

                response.raise_for_status()
                payload = response.json()

                if payload.get("status") == "error" or payload.get("code"):
                    raise RuntimeError(f"Twelve Data error: {payload}")

                return payload

            except (requests.RequestException, RuntimeError, ValueError) as exc:
                last_error = exc
                if attempt == self.config.retries - 1:
                    break
                wait = 2 ** attempt
                print(f"    [Retry {attempt + 1}/{self.config.retries - 1}] Error: {exc}. Retrying in {wait}s...")
                time.sleep(wait)

        raise RuntimeError(f"Request failed after retries: {last_error}")

    def fetch_time_series(self, interval: str) -> pd.DataFrame:
        start_dt, end_dt = resolve_date_range(self.config.years, self.config.start_date, self.config.end_date)
        print(f"    Resolved request range: {start_dt} -> {end_dt}")
        chunks = []
        current_start = start_dt

        if interval == "1day":
            step = relativedelta(years=max(1, self.config.years))
        elif interval == "4h":
            step = relativedelta(months=12)
        else:
            step = relativedelta(months=9)

        while current_start <= end_dt:
            # Inclusive intraday chunks. Date-only end dates are already expanded
            # to 23:59:59, so same-day 2h/4h bars are not cut at midnight.
            current_end = min(pd.Timestamp(current_start + step) - pd.Timedelta(seconds=1), end_dt)
            params = {
                "symbol": self.config.symbol,
                "interval": interval,
                "start_date": current_start.strftime("%Y-%m-%d %H:%M:%S"),
                "end_date": current_end.strftime("%Y-%m-%d %H:%M:%S"),
                "order": "asc",
                "outputsize": 5000,
                "format": "JSON",
                "adjust": self.config.adjust,
                "timezone": self.config.timezone,
            }
            if self.config.exchange:
                params["exchange"] = self.config.exchange

            payload = self._request("time_series", params)
            values = payload.get("values", [])

            if values:
                frame = pd.DataFrame(values)
                if "datetime" not in frame.columns:
                    raise RuntimeError(f"Twelve Data response missing datetime: {payload}")

                frame["datetime"] = pd.to_datetime(frame["datetime"], errors="coerce")
                frame = frame.dropna(subset=["datetime"])
                frame = frame.set_index("datetime").sort_index()

                for col in ("open", "high", "low", "close", "volume"):
                    if col in frame.columns:
                        frame[col] = pd.to_numeric(frame[col], errors="coerce")

                required_cols = ["open", "high", "low", "close"]
                missing_required = [col for col in required_cols if col not in frame.columns]
                if missing_required:
                    raise RuntimeError(
                        f"Twelve Data response missing OHLC columns {missing_required}"
                    )

                if "volume" not in frame.columns:
                    frame["volume"] = 0.0

                frame = frame[["open", "high", "low", "close", "volume"]].dropna()
                chunks.append(frame)

            current_start = current_end + pd.Timedelta(seconds=1)
            time.sleep(self.config.pause_sec)

        if not chunks:
            return pd.DataFrame()

        df = pd.concat(chunks).sort_index()
        df = df[~df.index.duplicated(keep="last")]
        df = df[(df.index >= start_dt) & (df.index <= end_dt)]
        return df.dropna()


def prepare_datasets(
    client: TwelveDataClient,
    selected_timeframes: List[str],
) -> Dict[str, pd.DataFrame]:
    datasets: Dict[str, pd.DataFrame] = {}

    load_intervals = list(selected_timeframes)
    if any(tf in INTRADAY_INTERVALS for tf in selected_timeframes) and "1day" not in load_intervals:
        load_intervals.append("1day")
        print("  Note: loading 1day as HTF helper only (not optimized unless explicitly selected).")

    for interval in load_intervals:
        print(f"  Loading {interval} data from Twelve Data...")
        df = client.fetch_time_series(interval)
        if df.empty:
            print(f"    No data returned for {interval}")
            continue
        datasets[interval] = df
        print(f"    {interval}: {len(df)} bars  {df.index[0]} -> {df.index[-1]}")

    if any(tf in INTRADAY_INTERVALS for tf in selected_timeframes) and "1day" not in datasets:
        raise ValueError("Daily data (1day) is required as HTF helper for intraday optimization but was not returned.")

    missing = [tf for tf in selected_timeframes if tf not in datasets]
    if missing:
        raise ValueError(f"Requested timeframes not returned: {missing}")

    return datasets


# ─────────────────────────────────────────────
# BACKTESTER
# ─────────────────────────────────────────────

def empty_backtest_result(df: pd.DataFrame, initial_capital: float) -> BacktestResult:
    last_bar_time = df.index[-1] if len(df) else pd.NaT
    last_close = safe_float(df["close"].iloc[-1]) if len(df) and "close" in df.columns else np.nan

    return {
        "net_profit_pct": 0.0,
        "net_profit_value": 0.0,
        "num_trades": 0,
        "num_winning": 0,
        "win_rate": 0.0,
        "max_drawdown_pct": 0.0,
        "profit_factor": 0.0,
        "sharpe": 0.0,
        "trades": [],
        "open_trade": None,
        "equity_curve": [],
        "last_bar_time": last_bar_time,
        "last_close": last_close,
        "final_equity": initial_capital,
    }


def backtest(
    df: pd.DataFrame,
    params: dict,
    htf_df: Optional[pd.DataFrame] = None,
    keep_open_trade: bool = True,
) -> BacktestResult:
    p = params
    initial_capital = p.get("initial_capital", 10000.0)
    position_pct = p.get("position_pct", 20.0) / 100.0
    commission_pct = p.get("commission_pct", 0.1) / 100.0

    if df.empty:
        return empty_backtest_result(df, initial_capital)

    open_ = df["open"].values
    high_ = df["high"].values
    low_ = df["low"].values
    close = df["close"].values
    n = len(df)

    ema_pull_v = ema(df["close"], p["ema_pull_len"]).values
    ema_mid_v = ema(df["close"], p["ema_mid_len"]).values
    ema_slow_v = ema(df["close"], p["ema_slow_len"]).values
    atr_v = atr(df, p["atr_len"]).values
    rsi_v = rsi(df["close"], p["rsi_len"]).values if p["use_rsi"] else None

    # FIX I1: avoid using current bar low in pullback test.
    pullback_low = df["low"].rolling(p["pullback_window"]).min().shift(1).values
    breakout_high = df["high"].rolling(p["breakout_lookback"]).max().shift(1).values

    if p["use_htf"] and htf_df is not None:
        htf_close_v, htf_ema_v = align_htf_filter(htf_df, df, p["htf_ema_len"])
    elif p["use_htf"]:
        htf_close_v = close
        htf_ema_v = ema(df["close"], p["htf_ema_len"]).values
    else:
        htf_close_v = None
        htf_ema_v = None

    warmup = max(
        p["ema_slow_len"],
        p["htf_ema_len"] if p["use_htf"] else 0,
        p["pullback_window"],
        p["breakout_lookback"],
        p["atr_len"],
        p["rsi_len"] if p["use_rsi"] else 0,
        50,
    )

    if n <= warmup + 5:
        return empty_backtest_result(df, initial_capital)

    equity = initial_capital
    equity_curve: List[float] = []
    trades: List[dict] = []
    open_trade: Optional[dict] = None

    in_position = False
    pending_entry = False
    pending_entry_atr = 0.0

    entry_price = 0.0
    entry_atr_val = 0.0
    r_dist_fixed = 0.0
    tp_fixed = 0.0
    position_size = 0.0
    remaining_entry_commission = 0.0
    position_bar = -1
    high_since = 0.0
    partial_taken = False
    last_exit_bar = np.nan
    current_trade_realized = 0.0
    current_trade_entry_time = None
    current_trade_partial_count = 0
    trade_initial_size = 0.0
    trade_return_pct_no_commission = 0.0

    for i in range(warmup, n):
        if pending_entry and not in_position:
            entry_price = open_[i]
            entry_atr_val = pending_entry_atr if pending_entry_atr > 0 else max(
                atr_v[i] if not np.isnan(atr_v[i]) else 0.0,
                _MIN_ATR_VALUE,
            )
            notional = equity * position_pct
            position_size = notional / entry_price if entry_price > 0 else 0.0
            remaining_entry_commission = notional * commission_pct

            in_position = position_size > 0
            position_bar = i
            high_since = high_[i]
            partial_taken = False
            current_trade_realized = 0.0
            current_trade_entry_time = df.index[i]
            current_trade_partial_count = 0
            trade_initial_size = position_size
            trade_return_pct_no_commission = 0.0
            pending_entry = False

            r_dist_fixed = max(entry_atr_val * p["sl_mult"], _MIN_ATR_VALUE)
            tp_fixed = entry_price + entry_atr_val * p["tp_mult"]

        if not in_position:
            c = close[i]

            valid = (
                not np.isnan(ema_pull_v[i])
                and not np.isnan(ema_mid_v[i])
                and not np.isnan(ema_slow_v[i])
                and c > ema_slow_v[i]
                and ema_mid_v[i] > ema_slow_v[i]
            )

            if valid and p["use_htf"]:
                valid = (
                    htf_close_v is not None
                    and htf_ema_v is not None
                    and not np.isnan(htf_close_v[i])
                    and not np.isnan(htf_ema_v[i])
                    and htf_close_v[i] > htf_ema_v[i]
                )

            if valid and p["use_rsi"]:
                valid = (
                    rsi_v is not None
                    and not np.isnan(rsi_v[i])
                    and p["rsi_min"] <= rsi_v[i] <= p["rsi_max"]
                )

            if valid and p["use_stretch_filter"]:
                atr_now = atr_v[i]
                valid = (
                    not np.isnan(atr_now)
                    and atr_now > 0
                    and (max(c - ema_pull_v[i], 0.0) / atr_now) <= p["max_stretch_atr"]
                )

            if valid:
                cooldown_ok = np.isnan(last_exit_bar) or ((i - last_exit_bar) > p["cooldown_bars"])
                if cooldown_ok:
                    use_pullback = p["entry_mode"] in ("Pullback Only", "Both")
                    use_breakout = p["entry_mode"] in ("Breakout Only", "Both")

                    pullback_ok = (
                        use_pullback
                        and not np.isnan(pullback_low[i])
                        and pullback_low[i] <= ema_pull_v[i]
                        and c > ema_pull_v[i]
                    )
                    breakout_ok = (
                        use_breakout
                        and not np.isnan(breakout_high[i])
                        and c > breakout_high[i]
                    )

                    if pullback_ok or breakout_ok:
                        pending_entry = True
                        pending_entry_atr = atr_v[i] if not np.isnan(atr_v[i]) else 0.0

        elif in_position:
            high_since = max(high_since, high_[i])
            atr_ref = entry_atr_val if entry_atr_val > 0 else max(
                atr_v[i] if not np.isnan(atr_v[i]) else _MIN_ATR_VALUE,
                _MIN_ATR_VALUE,
            )

            sl = entry_price - r_dist_fixed
            if p["use_be"] and close[i] >= entry_price + r_dist_fixed * p["be_r"]:
                sl = max(sl, entry_price)
            if partial_taken:
                sl = max(sl, entry_price)
            if p["use_chand"]:
                cur_atr = atr_v[i] if not np.isnan(atr_v[i]) else atr_ref
                chand_sl = high_since - cur_atr * p["chand_mult"]
                sl = max(sl, chand_sl)

            full_exit = False
            exit_price = None
            exit_reason = None

            if i != position_bar and open_[i] <= sl:
                exit_price = open_[i]
                exit_reason = "SL_GAP"
                full_exit = True
            elif i != position_bar and open_[i] >= tp_fixed:
                exit_price = open_[i]
                exit_reason = "TP_GAP"
                full_exit = True
            elif low_[i] <= sl:
                exit_price = sl
                exit_reason = "SL"
                full_exit = True
            else:
                if p["use_partial_tp"] and not partial_taken:
                    tp1 = entry_price + r_dist_fixed * p["partial_tp_r"]
                    if high_[i] >= tp1 and position_size > 0:
                        qty_before = position_size
                        exit_qty = qty_before * (p["partial_tp_pct"] / 100.0)
                        if exit_qty > 0:
                            entry_comm_alloc = remaining_entry_commission * (exit_qty / qty_before)
                            exit_commission = tp1 * exit_qty * commission_pct
                            pnl_partial = (
                                (tp1 - entry_price) * exit_qty
                                - entry_comm_alloc
                                - exit_commission
                            )
                            equity += pnl_partial
                            current_trade_realized += pnl_partial
                            current_trade_partial_count += 1
                            if trade_initial_size > 0:
                                trade_return_pct_no_commission += (exit_qty / trade_initial_size) * ((tp1 / entry_price - 1.0) * 100.0)
                            position_size -= exit_qty
                            remaining_entry_commission -= entry_comm_alloc
                            partial_taken = True

                if in_position and position_size > 0 and high_[i] >= tp_fixed:
                    exit_price = tp_fixed
                    exit_reason = "TP"
                    full_exit = True

            if not full_exit and in_position and p["use_time_stop"]:
                bars_in_trade = i - position_bar
                if bars_in_trade >= p["max_bars_in_trade"]:
                    exit_price = close[i]
                    exit_reason = "TIME"
                    full_exit = True

            if full_exit and in_position and position_size > 0:
                exit_qty = position_size
                entry_comm_alloc = remaining_entry_commission
                exit_commission = exit_price * exit_qty * commission_pct
                pnl_final = (
                    (exit_price - entry_price) * exit_qty
                    - entry_comm_alloc
                    - exit_commission
                )
                equity += pnl_final
                total_trade_pnl = current_trade_realized + pnl_final
                if trade_initial_size > 0:
                    remaining_fraction = exit_qty / trade_initial_size
                    trade_return_total_pct = trade_return_pct_no_commission + remaining_fraction * ((exit_price / entry_price - 1.0) * 100.0)
                else:
                    trade_return_total_pct = (exit_price / entry_price - 1.0) * 100.0

                trades.append(
                    {
                        "entry_time": current_trade_entry_time,
                        "entry_price": entry_price,
                        "exit_time": df.index[i],
                        "exit_price": exit_price,
                        "reason": exit_reason,
                        "pnl": total_trade_pnl,
                        "pnl_pct": (exit_price / entry_price - 1.0) * 100.0,
                        "trade_return_pct": trade_return_total_pct,
                        "partials": current_trade_partial_count,
                        "status": "CLOSED",
                    }
                )

                in_position = False
                position_size = 0.0
                remaining_entry_commission = 0.0
                entry_atr_val = 0.0
                r_dist_fixed = 0.0
                tp_fixed = 0.0
                partial_taken = False
                pending_entry = False
                last_exit_bar = i
                high_since = 0.0
                position_bar = -1
                current_trade_realized = 0.0
                current_trade_entry_time = None
                current_trade_partial_count = 0
                trade_initial_size = 0.0
                trade_return_pct_no_commission = 0.0

        # FIX I4: exactly one equity value per bar.
        if in_position and position_size > 0:
            liquidation_exit_comm = close[i] * position_size * commission_pct
            unrealized = (
                (close[i] - entry_price) * position_size
                - remaining_entry_commission
                - liquidation_exit_comm
            )
            equity_curve.append(equity + unrealized)
        else:
            equity_curve.append(equity)

    if in_position and position_size > 0:
        last_close = close[-1]
        liquidation_exit_comm = last_close * position_size * commission_pct
        unrealized = (
            (last_close - entry_price) * position_size
            - remaining_entry_commission
            - liquidation_exit_comm
        )
        if trade_initial_size > 0:
            remaining_fraction = position_size / trade_initial_size
            open_trade_return_pct = trade_return_pct_no_commission + remaining_fraction * ((last_close / entry_price - 1.0) * 100.0)
        else:
            open_trade_return_pct = (last_close / entry_price - 1.0) * 100.0

        cur_atr = atr_v[-1] if not np.isnan(atr_v[-1]) else entry_atr_val
        current_stop = entry_price - r_dist_fixed

        if p["use_be"] and last_close >= entry_price + r_dist_fixed * p["be_r"]:
            current_stop = max(current_stop, entry_price)
        if partial_taken:
            current_stop = max(current_stop, entry_price)
        if p["use_chand"]:
            chand_stop = high_since - cur_atr * p["chand_mult"]
            current_stop = max(current_stop, chand_stop)

        current_target = tp_fixed
        bars_in_trade = (n - 1) - position_bar if position_bar >= 0 else 0

        # Trade plan percentages should be measured from the entry price.
        # These are the original risk/reward percentages of the active trade.
        risk_to_stop_pct = ((current_stop / entry_price) - 1.0) * 100.0 if entry_price > 0 else np.nan
        reward_to_target_pct = ((current_target / entry_price) - 1.0) * 100.0 if entry_price > 0 else np.nan

        # Separate live distances from the current price. These are used only
        # for near-stop / near-target warnings and should not replace R/R.
        distance_to_stop_pct = ((current_stop / last_close) - 1.0) * 100.0 if last_close > 0 else np.nan
        distance_to_target_pct = ((current_target / last_close) - 1.0) * 100.0 if last_close > 0 else np.nan

        open_trade = {
            "entry_time": current_trade_entry_time,
            "entry_price": entry_price,
            "exit_time": pd.NaT,
            "exit_price": np.nan,
            "reason": "OPEN",
            "pnl": current_trade_realized + unrealized,
            "pnl_pct": (last_close / entry_price - 1.0) * 100.0,
            "trade_return_pct": open_trade_return_pct,
            "partials": current_trade_partial_count,
            "status": "OPEN",
            "current_price": last_close,
            "current_stop": current_stop,
            "current_target": current_target,
            "risk_to_stop_pct": risk_to_stop_pct,
            "reward_to_target_pct": reward_to_target_pct,
            "distance_to_stop_pct": distance_to_stop_pct,
            "distance_to_target_pct": distance_to_target_pct,
            "bars_in_trade": bars_in_trade,
            "action": "HOLD",
        }
        if not keep_open_trade:
            equity += current_trade_realized + unrealized
            trades.append(open_trade)
            open_trade = None

    closed_trades = trades
    num_trades = len(closed_trades)
    wins = [t for t in closed_trades if t["pnl"] > 0]
    losses = [t for t in closed_trades if t["pnl"] <= 0]

    gross_profit = sum(t["pnl"] for t in wins) if wins else 0.0
    gross_loss = abs(sum(t["pnl"] for t in losses)) if losses else 0.0

    if gross_loss > _PROFIT_FACTOR_FLOOR:
        profit_factor = gross_profit / gross_loss
    elif gross_profit > 0:
        profit_factor = float("inf")
    else:
        profit_factor = 0.0

    profit_factor_display = min(profit_factor, _PROFIT_FACTOR_DISPLAY_CAP) if math.isinf(profit_factor) else profit_factor

    eq = np.array(equity_curve, dtype=float)
    peak = np.maximum.accumulate(eq) if len(eq) else np.array([initial_capital])
    dd = (eq - peak) / np.where(peak > 0, peak, 1.0) if len(eq) else np.array([0.0])
    max_dd = abs(dd.min()) * 100.0 if len(dd) else 0.0

    if len(eq) > 1:
        rets = np.diff(eq) / np.where(eq[:-1] > 0, eq[:-1], 1.0)
        ann_factor = annualization_factor(df.index)
        sharpe = (rets.mean() / rets.std() * np.sqrt(ann_factor)) if rets.std() > 0 else 0.0
    else:
        sharpe = 0.0

    current_equity = eq[-1] if len(eq) else initial_capital

    return {
        "net_profit_pct": (current_equity / initial_capital - 1.0) * 100.0,
        "net_profit_value": current_equity - initial_capital,
        "num_trades": num_trades,
        "num_winning": len(wins),
        "win_rate": (len(wins) / num_trades * 100.0) if num_trades > 0 else 0.0,
        "max_drawdown_pct": max_dd,
        "profit_factor": profit_factor_display,
        "sharpe": sharpe,
        "trades": closed_trades,
        "open_trade": open_trade,
        "equity_curve": equity_curve,
        "last_bar_time": df.index[-1],
        "last_close": close[-1],
        "final_equity": current_equity,
        "pending_signal": {
            "status": "ENTRY_PENDING_NEXT_BAR",
            "signal_time": df.index[-1],
            "expected_entry": "next_bar_open",
            "pending_atr": pending_entry_atr,
        } if (pending_entry and not in_position) else None,
    }


# ─────────────────────────────────────────────
# OPTIMIZATION
# ─────────────────────────────────────────────

def create_objective(
    datasets: Dict[str, pd.DataFrame],
    min_trades: int,
    selected_timeframes: List[str],
):
    available_tfs = [tf for tf in selected_timeframes if tf in datasets]
    daily_htf = datasets.get("1day")

    def objective(trial: optuna.Trial) -> float:
        params = copy.deepcopy(DEFAULT_PARAMS)
        tf = trial.suggest_categorical("timeframe", available_tfs)
        df = datasets[tf]
        htf_data = daily_htf if tf in INTRADAY_INTERVALS else df

        params["use_htf"] = trial.suggest_categorical("use_htf", [True, False])
        params["htf_ema_len"] = trial.suggest_int("htf_ema_len", 100, 300)
        params["ema_pull_len"] = trial.suggest_int("ema_pull_len", 5, 25)
        params["ema_mid_len"] = trial.suggest_int("ema_mid_len", 15, 60)
        params["ema_slow_len"] = trial.suggest_int("ema_slow_len", 50, 220)

        # FIX I3: prune inverted EMA length hierarchy.
        if params["ema_pull_len"] >= params["ema_mid_len"]:
            raise optuna.exceptions.TrialPruned()
        if params["ema_mid_len"] >= params["ema_slow_len"]:
            raise optuna.exceptions.TrialPruned()

        params["entry_mode"] = trial.suggest_categorical(
            "entry_mode",
            ["Pullback Only", "Breakout Only", "Both"],
        )
        params["pullback_window"] = trial.suggest_int("pullback_window", 4, 20)
        params["breakout_lookback"] = trial.suggest_int("breakout_lookback", 10, 45)
        params["cooldown_bars"] = trial.suggest_int("cooldown_bars", 0, 10)

        params["use_rsi"] = trial.suggest_categorical("use_rsi", [True, False])
        if params["use_rsi"]:
            params["rsi_min"] = trial.suggest_int("rsi_min", 35, 55)
            params["rsi_max"] = trial.suggest_int("rsi_max", 60, 85)

        params["use_stretch_filter"] = trial.suggest_categorical("use_stretch_filter", [True, False])
        if params["use_stretch_filter"]:
            params["max_stretch_atr"] = trial.suggest_float("max_stretch_atr", 0.75, 3.0, step=0.25)

        params["atr_len"] = trial.suggest_int("atr_len", 7, 20)
        params["sl_mult"] = trial.suggest_float("sl_mult", 1.0, 4.0, step=0.25)
        params["tp_mult"] = trial.suggest_float("tp_mult", 2.0, 10.0, step=0.5)

        params["use_be"] = trial.suggest_categorical("use_be", [True, False])
        if params["use_be"]:
            params["be_r"] = trial.suggest_float("be_r", 0.5, 2.0, step=0.25)

        params["use_chand"] = trial.suggest_categorical("use_chand", [True, False])
        if params["use_chand"]:
            params["chand_mult"] = trial.suggest_float("chand_mult", 2.0, 6.0, step=0.25)

        params["use_partial_tp"] = trial.suggest_categorical("use_partial_tp", [True, False])
        if params["use_partial_tp"]:
            params["partial_tp_r"] = trial.suggest_float("partial_tp_r", 1.0, 3.0, step=0.25)
            params["partial_tp_pct"] = trial.suggest_int("partial_tp_pct", 20, 60)

        params["use_time_stop"] = trial.suggest_categorical("use_time_stop", [True, False])
        if params["use_time_stop"]:
            params["max_bars_in_trade"] = trial.suggest_int("max_bars_in_trade", 8, 60)

        result = backtest(df, params, htf_df=htf_data, keep_open_trade=True)
        net_pct = result["net_profit_pct"]
        win_rate = result["win_rate"]
        num_winning = result["num_winning"]
        num_trades = result["num_trades"]
        max_dd = result["max_drawdown_pct"]

        open_trade_bonus = 0.0
        if result["open_trade"] is not None and result["open_trade"]["pnl"] > 0:
            open_trade_bonus = min(result["open_trade"]["pnl_pct"], _OPEN_TRADE_BONUS_CAP) * _OPEN_TRADE_BONUS_WEIGHT

        if net_pct <= 0 or num_winning == 0:
            score = net_pct - 1000.0
        elif num_trades < min_trades:
            score = -1000.0 + num_trades
        else:
            score = (
                net_pct
                * (win_rate / 100.0)
                * math.log1p(num_winning)
                * math.log1p(max(num_trades, 1))
                + open_trade_bonus
            )

        trial.set_user_attr("timeframe", tf)
        trial.set_user_attr("num_trades", num_trades)
        trial.set_user_attr("num_winning", num_winning)
        trial.set_user_attr("win_rate", round(win_rate, 1))
        trial.set_user_attr("net_profit_pct", round(net_pct, 2))
        trial.set_user_attr("max_drawdown_pct", round(max_dd, 2))
        trial.set_user_attr("profit_factor", round(result["profit_factor"], 2))
        trial.set_user_attr("sharpe", round(result["sharpe"], 2))
        trial.set_user_attr("has_open_trade", result["open_trade"] is not None)
        return score

    return objective


def run_optimization(
    datasets: Dict[str, pd.DataFrame],
    min_trades: int,
    n_trials: int,
    seed: int,
    n_jobs: int,
    selected_timeframes: List[str],
) -> optuna.Study:
    sampler = TPESampler(
        seed=seed,
        n_startup_trials=min(50, max(10, n_trials // 5)),
    )
    study = optuna.create_study(
        direction="maximize",
        sampler=sampler,
        study_name="MultiTF_TwelveData_Optimizer",
    )

    print(f"\n  Starting Optimization ({n_trials} trials, jobs={n_jobs})")
    print(f"  Timeframes: {selected_timeframes}")
    print("  Objective : profit% x win_rate x log(winning_trades) x log(total_trades)")
    print("  " + "-" * 70)

    class ProgressCallback:
        def __init__(self):
            self.best = -float("inf")
            self.n = 0

        def __call__(self, study_obj, trial):
            self.n += 1
            if trial.value is not None and trial.value > self.best:
                self.best = trial.value
                a = trial.user_attrs
                print(
                    f"  [{self.n:4d}] New best! Score={trial.value:+.1f} | TF={a.get('timeframe','?')} | "
                    f"Profit={a.get('net_profit_pct',0):+.1f}% | WR={a.get('win_rate',0):.0f}% | "
                    f"Wins={a.get('num_winning',0)}/{a.get('num_trades',0)} | DD={a.get('max_drawdown_pct',0):.1f}%"
                )
            elif self.n % 100 == 0:
                print(f"  [{self.n:4d}] Searching... best={self.best:+.1f}")

    study.optimize(
        create_objective(datasets, min_trades, selected_timeframes),
        n_trials=n_trials,
        n_jobs=n_jobs,
        callbacks=[ProgressCallback()],
    )
    return study


# ─────────────────────────────────────────────
# REPORTING & EXPORT
# ─────────────────────────────────────────────

def apply_compound_capital(
    trades: pd.DataFrame,
    initial_capital: float = DEFAULT_INVESTMENT_AMOUNT,
) -> Tuple[pd.DataFrame, dict]:
    """
    Excel/report-only compounding calculation.

    Starts with the user-selected investment amount, uses 100% of current capital on every trade,
    compounds trade-by-trade, ignores position_pct and commission_pct.

    The trade return is taken from trade_return_pct, which is calculated
    without commissions and accounts for partial exits when present.
    """
    if trades.empty:
        return trades.copy(), {
            "compound_initial_capital": float(initial_capital),
            "compound_final_capital": initial_capital,
            "compound_net_profit_value": 0.0,
            "compound_net_profit_pct": 0.0,
            "compound_max_drawdown_pct": 0.0,
            "compound_num_winning": 0,
            "compound_win_rate": 0.0,
            "compound_uses_position_pct": False,
            "compound_uses_commission": False,
            "compound_includes_open_trade": False,
        }

    out = trades.copy()
    if "trade_return_pct" not in out.columns:
        out["trade_return_pct"] = out.get("pnl_pct", 0.0)

    capital = float(initial_capital)
    peak = capital
    max_dd = 0.0
    before_values = []
    after_values = []
    profit_values = []
    multipliers = []
    used_returns = []

    for _, row in out.iterrows():
        before = capital
        r_pct = safe_float(row.get("trade_return_pct", row.get("pnl_pct", 0.0)), 0.0)
        multiplier = 1.0 + (r_pct / 100.0)

        # Guard against impossible negative capital in extreme edge cases.
        if multiplier < 0:
            multiplier = 0.0

        after = before * multiplier
        peak = max(peak, after)
        dd = ((after - peak) / peak) * 100.0 if peak > 0 else 0.0
        max_dd = min(max_dd, dd)

        before_values.append(before)
        after_values.append(after)
        profit_values.append(after - before)
        multipliers.append(multiplier)
        used_returns.append(r_pct)
        capital = after

    out["compound_capital_before"] = before_values
    out["trade_return_pct_used_for_compound"] = used_returns
    out["compound_multiplier"] = multipliers
    out["compound_profit_value"] = profit_values
    out["compound_capital_after"] = after_values

    wins = sum(1 for r in used_returns if r > 0)
    total = len(used_returns)
    final_capital = capital
    metrics = {
        "compound_initial_capital": float(initial_capital),
        "compound_final_capital": final_capital,
        "compound_net_profit_value": final_capital - initial_capital,
        "compound_net_profit_pct": ((final_capital - initial_capital) / initial_capital) * 100.0,
        "compound_max_drawdown_pct": abs(max_dd),
        "compound_num_winning": wins,
        "compound_win_rate": (wins / total * 100.0) if total else 0.0,
        "compound_uses_position_pct": False,
        "compound_uses_commission": False,
        "compound_includes_open_trade": bool((out.get("status", pd.Series([], dtype=str)) == "OPEN").any()),
    }
    return out, metrics


def trade_diagnostics(trades: pd.DataFrame) -> dict:
    if trades.empty:
        return {
            "avg_win": 0.0,
            "avg_loss": 0.0,
            "avg_trade": 0.0,
            "median_trade": 0.0,
            "largest_win": 0.0,
            "largest_loss": 0.0,
            "expectancy_per_trade": 0.0,
            "closed_trades_count": 0,
            "open_trades_count": 0,
        }

    closed = trades[trades.get("status", "CLOSED") == "CLOSED"].copy()
    open_count = int((trades.get("status", pd.Series([], dtype=str)) == "OPEN").sum())

    if closed.empty:
        return {
            "avg_win": 0.0,
            "avg_loss": 0.0,
            "avg_trade": 0.0,
            "median_trade": 0.0,
            "largest_win": 0.0,
            "largest_loss": 0.0,
            "expectancy_per_trade": 0.0,
            "closed_trades_count": 0,
            "open_trades_count": open_count,
        }

    wins = closed[closed["pnl"] > 0]
    losses = closed[closed["pnl"] <= 0]

    avg_win = float(wins["pnl"].mean()) if not wins.empty else 0.0
    avg_loss = float(losses["pnl"].mean()) if not losses.empty else 0.0
    win_rate = len(wins) / len(closed) if len(closed) else 0.0
    loss_rate = 1.0 - win_rate

    return {
        "avg_win": avg_win,
        "avg_loss": avg_loss,
        "avg_trade": float(closed["pnl"].mean()),
        "median_trade": float(closed["pnl"].median()),
        "largest_win": float(closed["pnl"].max()),
        "largest_loss": float(closed["pnl"].min()),
        "expectancy_per_trade": (win_rate * avg_win) + (loss_rate * avg_loss),
        "closed_trades_count": int(len(closed)),
        "open_trades_count": open_count,
    }


def add_exit_reason_metrics(summary: dict, trades: pd.DataFrame) -> dict:
    if trades.empty or "reason" not in trades.columns:
        return summary

    reason_counts = trades["reason"].fillna("UNKNOWN").value_counts()
    for reason, count in reason_counts.items():
        summary[f"exit_reason_{reason}_count"] = int(count)

    return summary


def summarize_best(
    study: optuna.Study,
    datasets: Dict[str, pd.DataFrame],
    symbol: str,
    years: int,
    adjust: str,
    investment_amount: float = DEFAULT_INVESTMENT_AMOUNT,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Tuple[dict, pd.DataFrame, dict]:
    best = study.best_trial
    params = {**DEFAULT_PARAMS, **best.params}
    timeframe = best.params["timeframe"]
    df = datasets[timeframe]
    htf_df = datasets["1day"] if timeframe in INTRADAY_INTERVALS else df
    result = backtest(df, params, htf_df=htf_df, keep_open_trade=True)

    trades = pd.DataFrame(result["trades"])
    if result["open_trade"] is not None:
        trades = pd.concat([trades, pd.DataFrame([result["open_trade"]])], ignore_index=True)

    if not trades.empty:
        trades.insert(0, "trade_no", range(1, len(trades) + 1))
        trades.insert(1, "timeframe", timeframe)
        trades["entry_time"] = pd.to_datetime(trades["entry_time"])
        if "exit_time" in trades.columns:
            trades["exit_time"] = pd.to_datetime(trades["exit_time"], errors="coerce")

    trades, compound_metrics = apply_compound_capital(trades, investment_amount)
    diagnostics = trade_diagnostics(trades)

    summary = {
        "symbol": symbol,
        "years": years,
        "analysis_start": str(resolve_date_range(years, start_date, end_date)[0].date()),
        "analysis_end": str(resolve_date_range(years, start_date, end_date)[1].date()),
        "analysis_period": analysis_period_label(years, start_date, end_date),
        "investment_amount": float(investment_amount),
        "data_source": "Twelve Data",
        "adjust_mode": adjust,
        "atr_mode": ATR_MODE,
        "rsi_flat_mode": RSI_FLAT_MODE,
        "profit_calculation_for_excel": "compound_user_investment_full_capital_no_commission",
        "best_timeframe": timeframe,
        "score": best.value,
        "bars": len(df),
        "start": df.index[0],
        "end": df.index[-1],

        # Main Excel profit fields requested by the user:
        # start with $100, use 100% of current capital on each trade,
        # compound trade-by-trade, no commissions.
        "initial_capital": compound_metrics["compound_initial_capital"],
        "final_capital": compound_metrics["compound_final_capital"],
        "net_profit_value": compound_metrics["compound_net_profit_value"],
        "net_profit_pct": compound_metrics["compound_net_profit_pct"],
        "compound_max_drawdown_pct": compound_metrics["compound_max_drawdown_pct"],
        "compound_num_winning": compound_metrics["compound_num_winning"],
        "compound_win_rate": compound_metrics["compound_win_rate"],
        "compound_uses_position_pct": compound_metrics["compound_uses_position_pct"],
        "compound_uses_commission": compound_metrics["compound_uses_commission"],
        "compound_includes_open_trade": compound_metrics["compound_includes_open_trade"],

        # Original strategy/backtest metrics kept for reference.
        "strategy_initial_capital": params["initial_capital"],
        "strategy_position_pct": params.get("position_pct"),
        "strategy_commission_pct": params.get("commission_pct"),
        "strategy_net_profit_value": result["net_profit_value"],
        "strategy_net_profit_pct": result["net_profit_pct"],
        "num_trades": result["num_trades"],
        "num_winning": result["num_winning"],
        "win_rate": result["win_rate"],
        "max_drawdown_pct": result["max_drawdown_pct"],
        "profit_factor": result["profit_factor"],
        "sharpe": result["sharpe"],
        "has_open_trade": result["open_trade"] is not None,
        "last_bar_time": result["last_bar_time"],
        "last_close": result["last_close"],
        **diagnostics,
    }

    summary = add_exit_reason_metrics(summary, trades)
    return summary, trades, params



def _excel_blank(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return value


def _num(value, default=0.0):
    try:
        if value == "":
            return default
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def decision_reason(current_signal: dict) -> str:
    status = current_signal.get("current_status", "")
    action = current_signal.get("action", "")
    if status == "OPEN":
        pnl = _num(current_signal.get("current_pnl_pct"), 0.0)
        stop = current_signal.get("current_stop", "")
        target = current_signal.get("current_target", "")
        if pnl >= 0:
            tone = "الصفقة مفتوحة وربحانة حاليًا."
        else:
            tone = "الصفقة مفتوحة وخسرانة حاليًا، لكنها لم تصل إلى الستوب."
        return f"{tone} القرار الحالي {action}. تابع الستوب {stop} والهدف {target}."
    if status == "CLOSED":
        reason = current_signal.get("last_exit_reason", "")
        return f"آخر صفقة مغلقة بسبب {reason}. لا توجد صفقة مفتوحة حاليًا، القرار WAIT."
    if status == "ENTRY_PENDING_NEXT_BAR":
        return "ظهرت إشارة دخول على آخر شمعة. القرار: راقب افتتاح الشمعة القادمة قبل الدخول."
    return "لا توجد صفقة مفتوحة أو إشارة دخول حالية. القرار WAIT."


def compact_dashboard_df(summary: dict, current_signal: dict) -> pd.DataFrame:
    return pd.DataFrame([{
        "Symbol": summary.get("symbol"),
        "Mode": summary.get("mode", "OPTIMIZE"),
        "TF": summary.get("timeframe", summary.get("best_timeframe")),
        "Decision": current_signal.get("action"),
        "Status": current_signal.get("current_status"),
        "Current": current_signal.get("current_price"),
        "Entry": current_signal.get("last_entry_price"),
        "Stop": current_signal.get("current_stop"),
        "Target": current_signal.get("current_target"),
        "PnL %": current_signal.get("current_pnl_pct"),
        "Risk %": current_signal.get("risk_to_stop_pct"),
        "Reward %": current_signal.get("reward_to_target_pct"),
        "Final Capital": summary.get("final_capital"),
        "Profit %": summary.get("net_profit_pct"),
        "Trades": summary.get("num_trades"),
        "Win %": summary.get("win_rate"),
        "DD %": summary.get("max_drawdown_pct"),
        "PF": summary.get("profit_factor"),
        "Sharpe": summary.get("sharpe"),
        "Last Bar": summary.get("last_bar_time"),
    }])



SUMMARY_EXPLANATIONS = {
    "Decision": "القرار المختصر الآن: HOLD/WAIT/ENTER حسب آخر داتا.",
    "Status": "حالة الصفقة الحالية أو آخر صفقة.",
    "Reason": "شرح مختصر لسبب القرار.",
    "Symbol": "رمز السهم.",
    "Timeframe": "الفريم الذي فاز أو تم استخدامه.",
    "Analysis period": "فترة التحليل من تاريخ البداية إلى تاريخ النهاية.",
    "Investment amount": "المبلغ الذي اخترته للحساب المركب في التقرير.",
    "Last bar": "آخر شمعة وصلت من مزود البيانات.",
    "Current price": "آخر سعر متاح في الداتا.",
    "Entry price": "سعر دخول الصفقة الحالية أو آخر صفقة.",
    "Current stop": "وقف الخسارة الحالي للصفقة المفتوحة.",
    "Current target": "هدف الربح الحالي للصفقة المفتوحة.",
    "Current PnL %": "نسبة ربح/خسارة الصفقة الحالية.",
    "Risk to stop %": "المسافة من سعر الدخول إلى الستوب.",
    "Reward to target %": "المسافة من سعر الدخول إلى الهدف.",
    "Final capital": "كم أصبح مبلغ الاستثمار بعد الصفقات المركبة.",
    "Compound profit %": "نسبة الربح المركب على مبلغ الاستثمار.",
    "Trades": "عدد الصفقات المغلقة المستخدمة في الحساب.",
    "Win rate %": "نسبة الصفقات الرابحة.",
    "Max drawdown %": "أكبر هبوط من قمة رأس المال إلى قاع لاحق.",
    "Profit factor": "إجمالي الربح مقسوم على إجمالي الخسارة.",
    "Sharpe": "مقياس جودة العائد مقابل التذبذب.",
    "Params file": "ملف JSON المستخدم لتشغيل signal/replay.",
}


def compact_summary_df(summary: dict, current_signal: dict) -> pd.DataFrame:
    rows = [
        ("Decision", current_signal.get("action")),
        ("Status", current_signal.get("current_status")),
        ("Reason", decision_reason(current_signal)),
        ("Symbol", summary.get("symbol")),
        ("Timeframe", summary.get("timeframe", summary.get("best_timeframe"))),
        ("Analysis period", summary.get("analysis_period") or f"{summary.get('analysis_start', '')} -> {summary.get('analysis_end', '')}"),
        ("Investment amount", summary.get("investment_amount", summary.get("initial_capital"))),
        ("Last bar", summary.get("last_bar_time")),
        ("Current price", current_signal.get("current_price")),
        ("Entry price", current_signal.get("last_entry_price")),
        ("Current stop", current_signal.get("current_stop")),
        ("Current target", current_signal.get("current_target")),
        ("Current PnL %", current_signal.get("current_pnl_pct")),
        ("Risk to stop %", current_signal.get("risk_to_stop_pct")),
        ("Reward to target %", current_signal.get("reward_to_target_pct")),
        ("Final capital", summary.get("final_capital")),
        ("Compound profit %", summary.get("net_profit_pct")),
        ("Trades", summary.get("num_trades")),
        ("Win rate %", summary.get("win_rate")),
        ("Max drawdown %", summary.get("max_drawdown_pct")),
        ("Profit factor", summary.get("profit_factor")),
        ("Sharpe", summary.get("sharpe")),
        ("Params file", summary.get("params_file", "")),
    ]
    out = pd.DataFrame(rows, columns=["Item", "Value"])
    out["شرح"] = out["Item"].map(SUMMARY_EXPLANATIONS).fillna("")
    return out


def compact_trades_df(trades: pd.DataFrame) -> pd.DataFrame:
    if trades.empty:
        return pd.DataFrame(columns=["message"])

    cols = [
        "trade_no", "timeframe", "status", "entry_time", "entry_price",
        "exit_time", "exit_price", "reason", "trade_return_pct_used_for_compound",
        "compound_profit_value", "compound_capital_before", "compound_capital_after",
        "pnl_pct", "partials", "current_price", "current_stop", "current_target",
    ]
    existing = [c for c in cols if c in trades.columns]
    out = trades[existing].copy()
    rename = {
        "trade_no": "#",
        "timeframe": "TF",
        "status": "Status",
        "entry_time": "Entry Time",
        "entry_price": "Entry",
        "exit_time": "Exit Time",
        "exit_price": "Exit",
        "reason": "Reason",
        "trade_return_pct_used_for_compound": "Return %",
        "compound_profit_value": "Trade P/L $",
        "compound_capital_before": "Capital Before",
        "compound_capital_after": "Capital After",
        "pnl_pct": "Price PnL %",
        "partials": "Partials",
        "current_price": "Current",
        "current_stop": "Stop",
        "current_target": "Target",
    }
    return out.rename(columns=rename)




def report_guide_rows(report_type: str) -> list:
    common = [
        ["Action", "القرار الحالي", "HOLD = الصفقة مفتوحة وننتظر، WAIT = لا يوجد دخول حالي، ENTER_NEXT_BAR = إشارة دخول على افتتاح البار القادم"],
        ["Status", "حالة الصفقة", "OPEN = صفقة مفتوحة، CLOSED = آخر صفقة مغلقة، NO_POSITION = لا يوجد مركز"],
        ["Current", "السعر الحالي", "آخر سعر إغلاق/آخر سعر متاح في الداتا"],
        ["Stop", "وقف الخسارة الحالي", "السعر الذي تعتبر عنده الاستراتيجية أن الصفقة خرجت بخسارة أو حماية ربح"],
        ["Target", "الهدف الحالي", "سعر جني الربح المتوقع حسب إعدادات الاستراتيجية"],
        ["Risk %", "المسافة إلى الستوب", "كم يبعد الستوب عن سعر الدخول كنسبة مئوية"],
        ["Reward %", "المسافة إلى الهدف", "كم يبعد الهدف عن سعر الدخول كنسبة مئوية"],
        ["PnL %", "ربح/خسارة الصفقة", "النسبة الحالية أو النهائية للصفقة"],
        ["Entry vs current %", "فرق السعر عن الدخول", "يوضح هل السعر الحالي فوق أو تحت سعر الدخول"],
        ["Final Capital", "رأس المال النهائي", "كم أصبح مبلغ الاستثمار المحدد بعد تطبيق الصفقات المركبة"],
        ["Compound Profit %", "الربح المركب", "نسبة الربح/الخسارة من مبلغ الاستثمار المحدد بطريقة إعادة الاستثمار"],
        ["Strategy P/L", "ربح الاستراتيجية الأصلي", "نتيجة الباكتيست الأصلية حسب position_pct وcommission_pct داخل الكود"],
        ["Win Rate", "نسبة الصفقات الرابحة", "عدد الصفقات الرابحة ÷ إجمالي الصفقات"],
        ["Max Drawdown", "أكبر هبوط", "أكبر تراجع من قمة رأس المال إلى قاع بعدها"],
        ["Profit Factor", "عامل الربح", "إجمالي الربح ÷ إجمالي الخسارة؛ أعلى من 1 جيد"],
        ["Sharpe", "جودة العائد مقابل التذبذب", "كلما كان أعلى كان الأداء أكثر سلاسة، لكنه ليس ضمانًا"],
        ["Investment Amount", "مبلغ الاستثمار", "المبلغ الذي اخترته ليتم حساب الأرباح المركبة والدولار عليه"],
        ["Final Capital", "المبلغ النهائي", "كم أصبح مبلغ الاستثمار بعد تطبيق الصفقات"],
        ["Trade P/L $", "ربح/خسارة الصفقة بالدولار", "ربح أو خسارة الصفقة بالدولار حسب رأس المال المركب قبل الصفقة"],
        ["Capital Before", "رأس المال قبل الصفقة", "رأس المال المركب قبل تنفيذ الصفقة"],
        ["Capital After", "رأس المال بعد الصفقة", "رأس المال المركب بعد تنفيذ الصفقة"],
        ["Trades", "عدد الصفقات", "عدد الصفقات المغلقة المستخدمة في الحسابات"],
    ]

    signal = [
        ["Signal sheet", "ملخص القرار", "أهم شيت للمتابعة اليومية: يعطي القرار الحالي بدون تفاصيل زائدة"],
        ["RecentEvents sheet", "آخر الأحداث", "يعرض آخر صفقتين مغلقتين + الصفقة المفتوحة الحالية إن وجدت"],
        ["Reason", "سبب القرار", "شرح مختصر لماذا القرار HOLD أو WAIT"],
        ["Entry", "سعر الدخول", "سعر دخول الصفقة"],
        ["Exit", "سعر الخروج", "سعر خروج الصفقة المغلقة"],
    ]

    optimize = [
        ["Dashboard sheet", "لوحة مختصرة", "ملخص سريع لأهم نتيجة وحالة آخر صفقة ومبلغ الاستثمار النهائي"],
        ["Summary sheet", "ملخص رقمي", "أهم الأرقام فقط بدون كل التفاصيل القديمة"],
        ["Trades sheet", "سجل الصفقات", "كل الصفقات الناتجة من أفضل إعدادات"],
        ["TopTrials sheet", "أفضل التجارب", "أفضل 10 تجارب من Optuna للمقارنة"],
        ["Params sheet", "الإعدادات", "الإعدادات الفائزة التي تحفظها كخطة"],
    ]

    replay = [
        ["Dashboard sheet", "لوحة المتابعة", "ملخص تطبيق نفس JSON على الداتا الحالية"],
        ["TradePlan sheet", "خطة الصفقة الحالية", "الدخول والستوب والهدف والقرار بشكل عملي"],
        ["Trades sheet", "سجل الصفقات", "الصفقات بعد إعادة تشغيل نفس الإعدادات بدون optimization"],
        ["ParamsUsed sheet", "الإعدادات المستخدمة", "نفس الإعدادات المقروءة من ملف JSON"],
    ]

    scan = [
        ["StrategyDashboard sheet", "داشبورد كل الخطط", "صف لكل سهم/خطة مع status وaction وstop وtarget"],
        ["OpenTrades sheet", "الصفقات المفتوحة", "فلتر للخطط التي لديها صفقة مفتوحة الآن"],
        ["Errors sheet", "الأخطاء", "أي ملف JSON فشل أثناء الفحص يظهر هنا"],
    ]

    rows = [["Field / Column", "المعنى", "شرح بسيط"]]
    rows.extend(common)
    if report_type == "signal":
        rows.extend(signal)
    elif report_type == "optimize":
        rows.extend(optimize)
    elif report_type == "replay":
        rows.extend(replay)
    elif report_type == "scan":
        rows.extend(scan)
    return rows


def guide_df(report_type: str) -> pd.DataFrame:
    rows = report_guide_rows(report_type)
    return pd.DataFrame(rows[1:], columns=rows[0])


def add_guide_sheet(writer, report_type: str) -> None:
    guide_df(report_type).to_excel(writer, sheet_name="Guide", index=False)


def add_signal_explanations(signal_df: pd.DataFrame) -> pd.DataFrame:
    explanations = {
        "Decision": "القرار المختصر الآن",
        "Status": "حالة آخر صفقة",
        "Reason": "سبب القرار بشكل بسيط",
        "Symbol": "رمز السهم",
        "Timeframe": "الفريم المستخدم",
        "Current price": "آخر سعر متاح",
        "Entry price": "سعر دخول الصفقة المفتوحة",
        "Current stop": "الستوب الحالي",
        "Current target": "الهدف الحالي",
        "Current PnL %": "ربح/خسارة الصفقة الحالية",
        "Entry vs current %": "هل السعر الحالي فوق أو تحت الدخول",
        "Risk to stop %": "المسافة للستوب",
        "Reward to target %": "المسافة للهدف",
        "Compound final": "رأس المال المركب من 100$",
        "Compound profit %": "الربح المركب من 100$",
        "Trades": "عدد الصفقات",
        "Win rate": "نسبة الربح",
        "Max drawdown": "أكبر هبوط",
        "Profit factor": "عامل الربح",
        "Sharpe": "جودة العائد مقابل التذبذب",
    }
    out = signal_df.copy()
    key_col = "Metric" if "Metric" in out.columns else ("Item" if "Item" in out.columns else None)
    if key_col and "شرح" not in out.columns:
        out["شرح"] = out[key_col].map(explanations).fillna("")
    return out


def recent_signal_events_df(trades: pd.DataFrame) -> pd.DataFrame:
    """
    Signal-mode mini history:
      - If there is an OPEN trade now: show the last 2 closed trades + the current open trade.
      - If no open trade: show the last 2 trades only.

    This keeps Signal short while preventing the user from missing a trade
    that closed before a new one opened.
    """
    cols = [
        "Event", "Status", "Entry Time", "Entry", "Exit Time", "Exit",
        "Reason", "Current", "Stop", "Target", "Return %", "Compound Before",
        "Compound After", "Notes",
    ]

    if trades.empty:
        return pd.DataFrame(columns=cols)

    t = trades.copy()
    if "status" not in t.columns:
        t["status"] = "CLOSED"

    open_rows = t[t["status"].astype(str).eq("OPEN")]
    has_open = not open_rows.empty

    if has_open:
        current_open = open_rows.tail(1)
        closed_recent = t[~t["status"].astype(str).eq("OPEN")].tail(2)
        selected = pd.concat([closed_recent, current_open], ignore_index=True)
    else:
        selected = t.tail(2).copy()

    rows = []
    for _, row in selected.iterrows():
        status = str(row.get("status", "CLOSED"))
        is_open = status == "OPEN"
        event = "CURRENT OPEN" if is_open else "RECENT CLOSED"
        notes = "الصفقة المفتوحة الحالية" if is_open else "صفقة أُغلقت مؤخرًا"

        rows.append({
            "Event": event,
            "Status": status,
            "Entry Time": row.get("entry_time", ""),
            "Entry": row.get("entry_price", ""),
            "Exit Time": "" if is_open else row.get("exit_time", ""),
            "Exit": "" if is_open else row.get("exit_price", ""),
            "Reason": row.get("reason", "OPEN" if is_open else ""),
            "Current": row.get("current_price", ""),
            "Stop": row.get("current_stop", ""),
            "Target": row.get("current_target", ""),
            "Return %": row.get("trade_return_pct_used_for_compound", row.get("trade_return_pct", row.get("pnl_pct", ""))),
            "Compound Before": row.get("compound_capital_before", ""),
            "Compound After": row.get("compound_capital_after", ""),
            "Notes": notes,
        })

    return pd.DataFrame(rows, columns=cols)



def compact_params_df(params: dict) -> pd.DataFrame:
    keys = [
        "timeframe", "entry_mode", "use_htf", "ema_pull_len", "ema_mid_len", "ema_slow_len",
        "pullback_window", "breakout_lookback", "cooldown_bars",
        "use_rsi", "rsi_min", "rsi_max",
        "use_stretch_filter", "max_stretch_atr",
        "atr_len", "sl_mult", "tp_mult",
        "use_be", "be_r", "use_chand", "chand_mult",
        "use_partial_tp", "partial_tp_r", "partial_tp_pct",
        "use_time_stop", "max_bars_in_trade",
    ]
    rows = [(k, params.get(k, "")) for k in keys if k in params]
    return pd.DataFrame(rows, columns=["Setting", "Value"])


def compact_top_trials_df(top_trials_df: pd.DataFrame) -> pd.DataFrame:
    if top_trials_df.empty:
        return top_trials_df
    cols = [
        "rank", "score", "timeframe", "net_profit_pct", "win_rate", "num_winning",
        "num_trades", "max_drawdown_pct", "profit_factor", "sharpe",
    ]
    existing = [c for c in cols if c in top_trials_df.columns]
    out = top_trials_df[existing].copy()
    return out.rename(columns={
        "rank": "Rank",
        "score": "Score",
        "timeframe": "TF",
        "net_profit_pct": "Strategy Profit %",
        "win_rate": "Win %",
        "num_winning": "Wins",
        "num_trades": "Trades",
        "max_drawdown_pct": "DD %",
        "profit_factor": "PF",
        "sharpe": "Sharpe",
    })


def _format_ws(ws, title_color="D9EAF7"):
    header_fill = PatternFill("solid", fgColor=title_color)
    section_fill = PatternFill("solid", fgColor="EEF4F8")
    border = Border(bottom=Side(style="thin", color="D9E2EC"))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(name="Calibri", size=10)

    for cell in ws[1]:
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = 10
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            header = str(ws.cell(row=1, column=cell.column).value or "").lower()
            if any(token in header for token in ["price", "entry", "exit", "stop", "target", "capital", "current", "$100"]):
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "$#,##0.00"
            if any(token in header for token in ["%", "profit", "pnl", "risk", "reward", "dd", "win"]):
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
            if any(token in header for token in ["date", "time", "bar"]):
                cell.number_format = "yyyy-mm-dd hh:mm"

    if ws.title in ("Signal", "TradePlan", "Summary"):
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 44
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 22


def _finish_writer(writer):
    for sheet_name in writer.sheets:
        _format_ws(writer.sheets[sheet_name])


def export_signal_excel(path: str, summary: dict, current_signal: dict, trades: pd.DataFrame) -> None:
    signal_df = add_signal_explanations(compact_summary_df(summary, current_signal))
    recent_df = recent_signal_events_df(trades)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        signal_df.to_excel(writer, sheet_name="Signal", index=False)
        recent_df.to_excel(writer, sheet_name="RecentEvents", index=False)
        add_guide_sheet(writer, "signal")

        _finish_writer(writer)

        ws = writer.sheets["Signal"]
        ws.insert_rows(1, 2)
        ws.merge_cells("A1:B1")
        title = ws["A1"]
        title.value = f"Midas Trend Rider Signal - {summary.get('symbol')} {summary.get('timeframe')}"
        title.font = Font(name="Calibri", size=14, bold=True, color="1F2937")
        title.fill = PatternFill("solid", fgColor="EAF4EC")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A4"

        if "RecentEvents" in writer.sheets:
            ew = writer.sheets["RecentEvents"]
            ew.insert_rows(1, 2)
            ew.merge_cells("A1:N1")
            et = ew["A1"]
            et.value = "Recent Events - آخر صفقتين + الصفقة المفتوحة الحالية إن وجدت"
            et.font = Font(name="Calibri", size=13, bold=True, color="1F2937")
            et.fill = PatternFill("solid", fgColor="FFF7ED")
            et.alignment = Alignment(horizontal="center", vertical="center")
            ew.row_dimensions[1].height = 24
            ew.freeze_panes = "A4"



def export_excel(
    path: str,
    summary: dict,
    params: dict,
    trades: pd.DataFrame,
    top_trials_df: pd.DataFrame,
) -> None:
    fake_result = {"last_close": summary.get("last_close", ""), "pending_signal": None}
    current_signal = current_signal_from_result(fake_result, trades)

    dashboard_df = compact_dashboard_df({**summary, "mode": "OPTIMIZE"}, current_signal)
    summary_df = compact_summary_df({**summary, "mode": "OPTIMIZE"}, current_signal)
    trades_df = compact_trades_df(trades)
    params_df = compact_params_df(params)
    top_df = compact_top_trials_df(top_trials_df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        trades_df.to_excel(writer, sheet_name="Trades", index=False)
        top_df.to_excel(writer, sheet_name="TopTrials", index=False)
        params_df.to_excel(writer, sheet_name="Params", index=False)
        add_guide_sheet(writer, "optimize")
        _finish_writer(writer)


def build_top_trials(study: optuna.Study) -> pd.DataFrame:
    rows = []
    trials = [t for t in study.trials if t.value is not None]
    top_trials = sorted(trials, key=lambda t: t.value, reverse=True)[:10]

    for i, trial in enumerate(top_trials, 1):
        row = {
            "rank": i,
            "score": trial.value,
            "timeframe": trial.user_attrs.get("timeframe"),
            "net_profit_pct": trial.user_attrs.get("net_profit_pct"),
            "win_rate": trial.user_attrs.get("win_rate"),
            "num_winning": trial.user_attrs.get("num_winning"),
            "num_trades": trial.user_attrs.get("num_trades"),
            "max_drawdown_pct": trial.user_attrs.get("max_drawdown_pct"),
            "profit_factor": trial.user_attrs.get("profit_factor"),
            "sharpe": trial.user_attrs.get("sharpe"),
            "has_open_trade": trial.user_attrs.get("has_open_trade"),
        }
        for k, v in sorted(trial.params.items()):
            row[f"param_{k}"] = v
        rows.append(row)

    return pd.DataFrame(rows)


def print_report(summary: dict, params: dict, show_params: bool = True) -> None:
    print("\n" + "=" * 64)
    print(f"  Best Result for {summary['symbol']}")
    print("=" * 64)
    print(f"\n  Best timeframe   : {summary['best_timeframe']}")
    print(f"  Final Capital    : {summary.get('final_capital', 0.0):.2f} from investment {summary.get('investment_amount', summary.get('initial_capital', 0.0)):.2f}")
    print(f"  Compound Profit  : {summary['net_profit_pct']:+.2f}%")
    print(f"  Compound P/L $   : {summary['net_profit_value']:+.2f}")
    print(f"  Winning Trades   : {summary['num_winning']} / {summary['num_trades']}")
    print(f"  Win Rate         : {summary['win_rate']:.1f}%")
    print(f"  Max Drawdown     : {summary['max_drawdown_pct']:.2f}%")

    pf = summary["profit_factor"]
    pf_str = "∞" if pf >= _PROFIT_FACTOR_DISPLAY_CAP else f"{pf:.2f}"
    print(f"  Profit Factor    : {pf_str}")

    print(f"  Sharpe           : {summary['sharpe']:.2f}")
    print(f"  Strategy P/L     : {summary.get('strategy_net_profit_pct', 0.0):+.2f}% using strategy sizing")
    print(f"  Expectancy       : {summary.get('expectancy_per_trade', 0.0):+.2f} per trade")
    print(f"  ATR Mode         : {summary.get('atr_mode')}")
    print(f"  RSI Flat Mode    : {summary.get('rsi_flat_mode')}")

    if summary["has_open_trade"]:
        print(f"  Open trade now   : YES (last close {summary['last_close']:.4f} @ {summary['last_bar_time']})")
    else:
        print("  Open trade now   : NO")

    if show_params:
        print("\n  Best parameters:")
        for k, v in sorted(params.items()):
            print(f"    {k:28s} = {v}")



# ─────────────────────────────────────────────
# MODES: REPLAY / SIGNAL / SCAN
# ─────────────────────────────────────────────

def load_params_from_json(params_file: str) -> Tuple[dict, dict]:
    """
    Load a saved optimization/replay JSON.

    Expected structure:
      {"summary": {...}, "params": {...}}

    Returns:
      params  : merged DEFAULT_PARAMS + saved params
      summary : saved summary when available
    """
    with open(params_file, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if "params" not in payload:
        raise ValueError("JSON file does not contain a 'params' object.")

    params = {**DEFAULT_PARAMS, **payload["params"]}
    summary = payload.get("summary", {})
    return params, summary


def infer_symbol_from_json(params_file: str, summary: dict, fallback: Optional[str] = None) -> str:
    if fallback:
        return fallback.upper()

    symbol = summary.get("symbol") or summary.get("original_analysis_symbol")
    if symbol:
        return str(symbol).upper()

    name = os.path.basename(params_file)
    for suffix in ("_best_strategy_", "_replay_", "_signal_"):
        if suffix in name:
            return name.split(suffix)[0].upper()

    return os.path.splitext(name)[0].split("_")[0].upper()


def get_replay_timeframe(params: dict, requested_timeframes: Optional[List[str]] = None) -> str:
    tf = params.get("timeframe")
    if tf in SUPPORTED_INTERVALS:
        return tf

    if requested_timeframes:
        return requested_timeframes[0]

    return "1day"


def prepare_trades_dataframe(
    result: dict,
    timeframe: str,
    investment_amount: float = DEFAULT_INVESTMENT_AMOUNT,
) -> pd.DataFrame:
    trades = pd.DataFrame(result["trades"])
    if result.get("open_trade") is not None:
        trades = pd.concat([trades, pd.DataFrame([result["open_trade"]])], ignore_index=True)

    if not trades.empty:
        trades.insert(0, "trade_no", range(1, len(trades) + 1))
        trades.insert(1, "timeframe", timeframe)
        trades["entry_time"] = pd.to_datetime(trades["entry_time"])
        if "exit_time" in trades.columns:
            trades["exit_time"] = pd.to_datetime(trades["exit_time"], errors="coerce")

    trades, _compound_metrics = apply_compound_capital(trades, investment_amount)
    return trades


def current_signal_from_result(result: dict, trades: pd.DataFrame) -> dict:
    pending = result.get("pending_signal")

    if not trades.empty:
        last = trades.iloc[-1]
        status = str(last.get("status", "CLOSED"))

        if status == "OPEN":
            entry_price = safe_float(last.get("entry_price"), np.nan)
            current_price = safe_float(last.get("current_price", result.get("last_close")), np.nan)
            current_stop = safe_float(last.get("current_stop"), np.nan)
            current_target = safe_float(last.get("current_target"), np.nan)

            current_pnl_pct = safe_float(last.get("pnl_pct"), np.nan)

            # R/R shown to the user is calculated from entry price.
            risk_to_stop_pct = (
                (current_stop / entry_price - 1.0) * 100.0
                if entry_price and entry_price > 0 and not np.isnan(current_stop)
                else safe_float(last.get("risk_to_stop_pct"), np.nan)
            )
            reward_to_target_pct = (
                (current_target / entry_price - 1.0) * 100.0
                if entry_price and entry_price > 0 and not np.isnan(current_target)
                else safe_float(last.get("reward_to_target_pct"), np.nan)
            )

            # Live distances are calculated from current price and used for warnings only.
            distance_to_stop_pct = (
                (current_stop / current_price - 1.0) * 100.0
                if current_price and current_price > 0 and not np.isnan(current_stop)
                else safe_float(last.get("distance_to_stop_pct"), np.nan)
            )
            distance_to_target_pct = (
                (current_target / current_price - 1.0) * 100.0
                if current_price and current_price > 0 and not np.isnan(current_target)
                else safe_float(last.get("distance_to_target_pct"), np.nan)
            )

            entry_vs_current_pct = (
                (current_price / entry_price - 1.0) * 100.0
                if entry_price and entry_price > 0 and not np.isnan(current_price)
                else np.nan
            )

            # Smarter action labels:
            # HOLD_PROFIT              = open and above entry
            # HOLD_LOSING_ABOVE_STOP   = open, losing, but still above stop
            # HOLD_NEAR_STOP           = open and close to stop
            # HOLD_NEAR_TARGET         = open and close to target
            # HOLD                     = fallback
            action = "HOLD"
            reason = "الصفقة مفتوحة ولم تضرب الستوب أو الهدف بعد."

            near_stop_threshold = 2.0      # percentage points to stop
            near_target_threshold = 2.0    # percentage points to target

            distance_to_stop_abs = abs(distance_to_stop_pct) if not np.isnan(distance_to_stop_pct) else np.nan
            distance_to_target_abs = abs(distance_to_target_pct) if not np.isnan(distance_to_target_pct) else np.nan

            if not np.isnan(distance_to_stop_abs) and distance_to_stop_abs <= near_stop_threshold:
                action = "HOLD_NEAR_STOP"
                reason = "الصفقة مفتوحة لكنها قريبة من الستوب. لم يحصل خروج بعد حسب الاستراتيجية."
            elif not np.isnan(distance_to_target_abs) and distance_to_target_abs <= near_target_threshold:
                action = "HOLD_NEAR_TARGET"
                reason = "الصفقة مفتوحة وقريبة من الهدف. لم يتحقق الهدف بعد حسب الاستراتيجية."
            elif not np.isnan(entry_vs_current_pct) and entry_vs_current_pct >= 0:
                action = "HOLD_PROFIT"
                reason = "الصفقة مفتوحة وربحانة حاليًا. الاستراتيجية تقول الاستمرار."
            elif not np.isnan(entry_vs_current_pct) and entry_vs_current_pct < 0:
                action = "HOLD_LOSING_ABOVE_STOP"
                reason = "الصفقة مفتوحة وخسرانة حاليًا، لكنها ما زالت فوق الستوب. لا يوجد خروج حسب الاستراتيجية."

            return {
                "action": action,
                "decision_reason": reason,
                "current_status": "OPEN",
                "last_entry_time": last.get("entry_time", ""),
                "last_entry_price": entry_price,
                "last_exit_time": "",
                "last_exit_price": "",
                "last_exit_reason": "OPEN",
                "current_price": current_price,
                "current_stop": current_stop,
                "current_target": current_target,
                "risk_to_stop_pct": risk_to_stop_pct,
                "reward_to_target_pct": reward_to_target_pct,
                "distance_to_stop_pct": distance_to_stop_pct,
                "distance_to_target_pct": distance_to_target_pct,
                "entry_vs_current_pct": entry_vs_current_pct,
                "bars_in_trade": last.get("bars_in_trade", ""),
                "current_pnl": last.get("pnl", ""),
                "current_pnl_pct": current_pnl_pct,
                "trade_return_pct": last.get("trade_return_pct", last.get("pnl_pct", "")),
                "notes": reason,
            }

        return {
            "action": "WAIT_AFTER_EXIT",
            "decision_reason": "آخر صفقة مغلقة. لا توجد صفقة مفتوحة حاليًا حسب نفس الإعدادات.",
            "current_status": "CLOSED",
            "last_entry_time": last.get("entry_time", ""),
            "last_entry_price": last.get("entry_price", ""),
            "last_exit_time": last.get("exit_time", ""),
            "last_exit_price": last.get("exit_price", ""),
            "last_exit_reason": last.get("reason", ""),
            "current_price": result.get("last_close", ""),
            "current_stop": "",
            "current_target": "",
            "risk_to_stop_pct": "",
            "reward_to_target_pct": "",
            "entry_vs_current_pct": "",
            "bars_in_trade": "",
            "current_pnl": last.get("pnl", ""),
            "current_pnl_pct": last.get("pnl_pct", ""),
            "trade_return_pct": last.get("trade_return_pct", last.get("pnl_pct", "")),
            "notes": "آخر صفقة سكرت. انتظر إشارة جديدة.",
        }

    if pending:
        return {
            "action": "ENTER_NEXT_BAR_CHECK_OPEN",
            "decision_reason": "ظهرت إشارة دخول على آخر بار. الدخول النظري يكون على افتتاح البار القادم.",
            "current_status": "ENTRY_PENDING_NEXT_BAR",
            "last_entry_time": "",
            "last_entry_price": "",
            "last_exit_time": "",
            "last_exit_price": "",
            "last_exit_reason": "",
            "current_price": result.get("last_close", ""),
            "current_stop": "",
            "current_target": "",
            "risk_to_stop_pct": "",
            "reward_to_target_pct": "",
            "entry_vs_current_pct": "",
            "bars_in_trade": "",
            "current_pnl": "",
            "current_pnl_pct": "",
            "trade_return_pct": "",
            "signal_time": pending.get("signal_time", ""),
            "notes": "إشارة دخول جديدة؛ راقب افتتاح البار القادم.",
        }

    return {
        "action": "WAIT",
        "decision_reason": "لا توجد صفقة مفتوحة ولا إشارة دخول حالية حسب الإعدادات.",
        "current_status": "NO_POSITION",
        "last_entry_time": "",
        "last_entry_price": "",
        "last_exit_time": "",
        "last_exit_price": "",
        "last_exit_reason": "",
        "current_price": result.get("last_close", ""),
        "current_stop": "",
        "current_target": "",
        "risk_to_stop_pct": "",
        "reward_to_target_pct": "",
        "entry_vs_current_pct": "",
        "bars_in_trade": "",
        "current_pnl": "",
        "current_pnl_pct": "",
        "trade_return_pct": "",
        "notes": "لا يوجد مركز حاليًا.",
    }


def build_summary_from_result(
    mode: str,
    symbol: str,
    years: int,
    adjust: str,
    params_file: str,
    timeframe: str,
    df: pd.DataFrame,
    result: dict,
    trades: pd.DataFrame,
    params: dict,
    current_signal: dict,
    old_summary: Optional[dict] = None,
    investment_amount: float = DEFAULT_INVESTMENT_AMOUNT,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    old_summary = old_summary or {}
    diagnostics = trade_diagnostics(trades)
    _trades_for_compound, compound_metrics = apply_compound_capital(trades, investment_amount)

    summary = {
        "mode": mode.upper(),
        "symbol": symbol,
        "years": years,
        "analysis_start": str(resolve_date_range(years, start_date, end_date)[0].date()),
        "analysis_end": str(resolve_date_range(years, start_date, end_date)[1].date()),
        "analysis_period": analysis_period_label(years, start_date, end_date),
        "investment_amount": float(investment_amount),
        "data_source": "Twelve Data",
        "adjust_mode": adjust,
        "atr_mode": ATR_MODE,
        "rsi_flat_mode": RSI_FLAT_MODE,
        "profit_calculation_for_excel": "compound_user_investment_full_capital_no_commission",
        "params_file": params_file,
        "original_analysis_symbol": old_summary.get("symbol", ""),
        "original_analysis_timeframe": old_summary.get("best_timeframe", old_summary.get("timeframe", "")),
        "timeframe": timeframe,
        "bars": len(df),
        "start": df.index[0],
        "end": df.index[-1],

        "initial_capital": compound_metrics["compound_initial_capital"],
        "final_capital": compound_metrics["compound_final_capital"],
        "net_profit_value": compound_metrics["compound_net_profit_value"],
        "net_profit_pct": compound_metrics["compound_net_profit_pct"],
        "compound_max_drawdown_pct": compound_metrics["compound_max_drawdown_pct"],
        "compound_num_winning": compound_metrics["compound_num_winning"],
        "compound_win_rate": compound_metrics["compound_win_rate"],
        "compound_uses_position_pct": False,
        "compound_uses_commission": False,
        "compound_includes_open_trade": compound_metrics["compound_includes_open_trade"],

        "strategy_initial_capital": params.get("initial_capital"),
        "strategy_position_pct": params.get("position_pct"),
        "strategy_commission_pct": params.get("commission_pct"),
        "strategy_net_profit_value": result["net_profit_value"],
        "strategy_net_profit_pct": result["net_profit_pct"],
        "num_trades": result["num_trades"],
        "num_winning": result["num_winning"],
        "win_rate": result["win_rate"],
        "max_drawdown_pct": result["max_drawdown_pct"],
        "profit_factor": result["profit_factor"],
        "sharpe": result["sharpe"],
        "has_open_trade": result.get("open_trade") is not None,
        "last_bar_time": result["last_bar_time"],
        "last_close": result["last_close"],
        **diagnostics,
        **{f"signal_{k}": v for k, v in current_signal.items()},
    }

    return add_exit_reason_metrics(summary, trades)


def export_replay_excel(
    path: str,
    summary: dict,
    params: dict,
    trades: pd.DataFrame,
    current_signal: dict,
) -> None:
    dashboard_df = compact_dashboard_df(summary, current_signal)
    summary_df = compact_summary_df(summary, current_signal)
    trades_df = compact_trades_df(trades)
    params_df = compact_params_df(params)
    trade_plan_df = pd.DataFrame([{
        "Symbol": summary.get("symbol"),
        "Timeframe": summary.get("timeframe"),
        "Action": current_signal.get("action"),
        "Status": current_signal.get("current_status"),
        "Entry": current_signal.get("last_entry_price"),
        "Current": current_signal.get("current_price"),
        "Stop": current_signal.get("current_stop"),
        "Target": current_signal.get("current_target"),
        "Risk %": current_signal.get("risk_to_stop_pct"),
        "Reward %": current_signal.get("reward_to_target_pct"),
        "Bars": current_signal.get("bars_in_trade"),
        "Reason": decision_reason(current_signal),
    }])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
        trade_plan_df.to_excel(writer, sheet_name="TradePlan", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        trades_df.to_excel(writer, sheet_name="Trades", index=False)
        params_df.to_excel(writer, sheet_name="Params", index=False)
        add_guide_sheet(writer, "replay")
        _finish_writer(writer)


def run_params_mode(args, mode: str) -> None:
    params_file = args.params_file or args.replay
    if not params_file:
        raise SystemExit(f"--mode {mode} requires --params-file or --replay.")

    params, old_summary = load_params_from_json(params_file)
    requested_tfs = parse_timeframes(args.timeframes)
    replay_tf = get_replay_timeframe(params, requested_tfs)
    selected_timeframes = [replay_tf]

    symbol = infer_symbol_from_json(params_file, old_summary, args.ticker)

    cfg = DownloadConfig(
        api_key=args.api_key or os.environ.get("TWELVEDATA_API_KEY") or DEFAULT_TWELVEDATA_API_KEY,
        symbol=symbol,
        years=args.years,
        start_date=args.start_date,
        end_date=args.end_date,
        adjust=args.adjust,
        exchange=args.exchange,
        pause_sec=max(float(getattr(args, "pause_sec", 0.35) or 0.35), 0.0),
    )

    if not cfg.api_key:
        raise SystemExit("Missing API key. Use --api-key or set TWELVEDATA_API_KEY.")

    client = TwelveDataClient(cfg)

    print(f"\n{mode.capitalize()} mode for {cfg.symbol}")
    print(f"Params file : {params_file}")
    print(f"Timeframe   : {replay_tf}")
    print("Optimization: SKIPPED")

    datasets = prepare_datasets(client, selected_timeframes)
    df = datasets[replay_tf]
    htf_df = datasets["1day"] if replay_tf in INTRADAY_INTERVALS else df

    params["timeframe"] = replay_tf
    result = backtest(df, params, htf_df=htf_df, keep_open_trade=True)

    trades = prepare_trades_dataframe(result, replay_tf, investment_amount=args.investment_amount)
    current_signal = current_signal_from_result(result, trades)
    summary = build_summary_from_result(
        mode=mode,
        symbol=cfg.symbol,
        years=args.years,
        adjust=args.adjust,
        params_file=params_file,
        timeframe=replay_tf,
        df=df,
        result=result,
        trades=trades,
        params=params,
        current_signal=current_signal,
        old_summary=old_summary,
        investment_amount=args.investment_amount,
        start_date=args.start_date,
        end_date=args.end_date,
    )

    os.makedirs(args.output, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "signal" if mode == "signal" else "replay"
    excel_path = os.path.join(args.output, f"{cfg.symbol}_{suffix}_{timestamp}.xlsx")
    json_path = os.path.join(args.output, f"{cfg.symbol}_{suffix}_{timestamp}.json")

    if mode == "signal":
        export_signal_excel(excel_path, summary, current_signal, trades)
    else:
        export_replay_excel(excel_path, summary, params, trades, current_signal)

    if not args.no_json:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(
                {"summary": summary, "params": params, "current_signal": current_signal, "recent_events": recent_signal_events_df(trades).to_dict(orient="records")},
                f,
                indent=2,
                default=str,
            )

    print(f"\n{mode.capitalize()} result")
    print(f"  Status           : {current_signal.get('current_status')}")
    print(f"  Action           : {current_signal.get('action')}")
    print(f"  Reason           : {current_signal.get('decision_reason', current_signal.get('notes', ''))}")
    print(f"  Compound Final $ : {summary.get('final_capital'):.2f}")
    print(f"  Compound Profit  : {summary.get('net_profit_pct'):+.2f}%")
    print(f"  Strategy P/L     : {summary.get('strategy_net_profit_pct'):+.2f}%")
    print(f"  Trades           : {result['num_winning']} / {result['num_trades']} wins")

    if mode == "signal":
        recent_events = recent_signal_events_df(trades)
        if not recent_events.empty:
            print("\nRecent events:")
            for _, ev in recent_events.iterrows():
                if ev.get("Status") == "OPEN":
                    print(f"  - OPEN | entry {ev.get('Entry Time')} @ {ev.get('Entry')} | current {ev.get('Current')} | return {ev.get('Return %')}%")
                else:
                    print(f"  - CLOSED | entry {ev.get('Entry Time')} -> exit {ev.get('Exit Time')} | {ev.get('Reason')} | return {ev.get('Return %')}%")

    if current_signal["current_status"] == "OPEN":
        print(f"  Last Entry       : {current_signal['last_entry_time']} @ {current_signal['last_entry_price']}")
        print(f"  Current Price    : {current_signal['current_price']}")
        print(f"  Current Stop     : {current_signal['current_stop']}")
        print(f"  Current Target   : {current_signal['current_target']}")
        print(f"  Current PnL      : {current_signal['current_pnl']} ({current_signal['current_pnl_pct']}%)")
    elif current_signal["current_status"] == "CLOSED":
        print(f"  Last Entry       : {current_signal['last_entry_time']} @ {current_signal['last_entry_price']}")
        print(f"  Last Exit        : {current_signal['last_exit_time']} @ {current_signal['last_exit_price']}")
        print(f"  Exit Reason      : {current_signal['last_exit_reason']}")
    else:
        print(f"  Notes            : {current_signal.get('notes')}")

    print("\nSaved files:")
    print(f"  Excel : {excel_path}")
    if not args.no_json:
        print(f"  JSON  : {json_path}")


def run_replay_mode(args) -> None:
    run_params_mode(args, mode="replay")


def run_signal_mode(args) -> None:
    run_params_mode(args, mode="signal")


def run_scan_mode(args) -> None:
    import glob

    pattern = args.scan_glob
    if not pattern:
        if args.params_file:
            pattern = args.params_file
        elif args.replay:
            pattern = args.replay
        else:
            pattern = os.path.join(args.output, "*_best_strategy_*.json")

    files = []
    for part in str(pattern).split(","):
        part = part.strip()
        if not part:
            continue
        matches = glob.glob(part)
        files.extend(matches if matches else [part])

    files = [f for f in files if os.path.exists(f)]
    if not files:
        raise SystemExit(f"No JSON files found for scan pattern: {pattern}")

    api_key = args.api_key or os.environ.get("TWELVEDATA_API_KEY") or DEFAULT_TWELVEDATA_API_KEY
    if not api_key:
        raise SystemExit("Missing API key. Use --api-key or set TWELVEDATA_API_KEY.")

    rows = []
    details = {}

    print(f"\nScan mode")
    print(f"JSON files : {len(files)}")
    print("Optimization: SKIPPED")

    for params_file in files:
        try:
            params, old_summary = load_params_from_json(params_file)
            symbol = infer_symbol_from_json(params_file, old_summary, None)
            tf = get_replay_timeframe(params, None)
            selected_timeframes = [tf]

            cfg = DownloadConfig(
                api_key=api_key,
                symbol=symbol,
                years=args.years,
                start_date=args.start_date,
                end_date=args.end_date,
                adjust=args.adjust,
                exchange=args.exchange,
                pause_sec=max(float(getattr(args, "pause_sec", 0.35) or 0.35), 0.0),
            )
            client = TwelveDataClient(cfg)

            print(f"\n  Scanning {symbol} | TF={tf}")
            datasets = prepare_datasets(client, selected_timeframes)
            df = datasets[tf]
            htf_df = datasets["1day"] if tf in INTRADAY_INTERVALS else df

            params["timeframe"] = tf
            result = backtest(df, params, htf_df=htf_df, keep_open_trade=True)
            trades = prepare_trades_dataframe(result, tf, investment_amount=args.investment_amount)
            current_signal = current_signal_from_result(result, trades)
            summary = build_summary_from_result(
                mode="scan",
                symbol=symbol,
                years=args.years,
                adjust=args.adjust,
                params_file=params_file,
                timeframe=tf,
                df=df,
                result=result,
                trades=trades,
                params=params,
                current_signal=current_signal,
                old_summary=old_summary,
                investment_amount=args.investment_amount,
                start_date=args.start_date,
                end_date=args.end_date,
            )

            rows.append({
                "symbol": symbol,
                "timeframe": tf,
                "status": current_signal.get("current_status"),
                "action": current_signal.get("action"),
                "entry_time": current_signal.get("last_entry_time"),
                "entry_price": current_signal.get("last_entry_price"),
                "current_price": current_signal.get("current_price"),
                "current_stop": current_signal.get("current_stop"),
                "current_target": current_signal.get("current_target"),
                "risk_to_stop_pct": current_signal.get("risk_to_stop_pct"),
                "reward_to_target_pct": current_signal.get("reward_to_target_pct"),
                "current_pnl_pct": current_signal.get("current_pnl_pct"),
                "compound_final_capital": summary.get("final_capital"),
                "compound_profit_pct": summary.get("net_profit_pct"),
                "strategy_profit_pct": summary.get("strategy_net_profit_pct"),
                "num_trades": summary.get("num_trades"),
                "win_rate": summary.get("win_rate"),
                "max_drawdown_pct": summary.get("max_drawdown_pct"),
                "profit_factor": summary.get("profit_factor"),
                "sharpe": summary.get("sharpe"),
                "last_bar_time": summary.get("last_bar_time"),
                "params_file": params_file,
            })
            details[symbol] = {
                "summary": summary,
                "trades": trades,
                "current_signal": current_signal,
                "params": params,
            }

        except Exception as exc:
            rows.append({
                "symbol": os.path.basename(params_file).split("_")[0],
                "timeframe": "",
                "status": "ERROR",
                "action": "CHECK_ERROR",
                "notes": str(exc),
                "params_file": params_file,
            })

    dashboard = pd.DataFrame(rows)
    os.makedirs(args.output, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(args.output, f"strategy_dashboard_{timestamp}.xlsx")
    json_path = os.path.join(args.output, f"strategy_dashboard_{timestamp}.json")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)

        open_rows = dashboard[dashboard["status"].eq("OPEN")] if "status" in dashboard.columns else pd.DataFrame()
        open_rows.to_excel(writer, sheet_name="OpenTrades", index=False)

        errors = dashboard[dashboard["status"].eq("ERROR")] if "status" in dashboard.columns else pd.DataFrame()
        if not errors.empty:
            errors.to_excel(writer, sheet_name="Errors", index=False)

        add_guide_sheet(writer, "scan")

        _finish_writer(writer)

    if not args.no_json:
        safe_details = {
            symbol: {
                "summary": info["summary"],
                "current_signal": info["current_signal"],
                "params": info["params"],
            }
            for symbol, info in details.items()
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"dashboard": rows, "details": safe_details}, f, indent=2, default=str)

    print("\nDashboard")
    if not dashboard.empty:
        print(dashboard[["symbol", "timeframe", "status", "action", "current_price", "current_stop", "current_target", "current_pnl_pct"]].to_string(index=False))

    print("\nSaved files:")
    print(f"  Excel : {excel_path}")
    if not args.no_json:
        print(f"  JSON  : {json_path}")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Midas Trend Rider: optimize, replay, signal, and scan modes")
    parser.add_argument("ticker", nargs="?", default=None, help="Ticker symbol, e.g. IREN. Not required for scan if JSON files contain symbols.")
    parser.add_argument(
        "--mode",
        type=str,
        default=None,
        choices=["optimize", "replay", "signal", "scan"],
        help="Run mode. Defaults to optimize unless --replay/--params-file is provided.",
    )
    parser.add_argument("--replay", type=str, default=None, help="Backward-compatible alias for --params-file in replay/signal mode.")
    parser.add_argument("--params-file", type=str, default=None, help="Saved JSON file with params for replay/signal, or comma/glob list for scan.")
    parser.add_argument("--scan-glob", type=str, default=None, help="Glob or comma-separated JSON list for --mode scan, e.g. '/content/results/*_best_strategy_*.json'.")
    parser.add_argument("--years", type=int, default=1, help="Years of history to request. Used only if --start-date is not provided.")
    parser.add_argument("--start-date", type=str, default=None, help="Analysis start date YYYY-MM-DD. Overrides --years when provided.")
    parser.add_argument("--end-date", type=str, default=None, help="Analysis end date YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--investment-amount", type=float, default=DEFAULT_INVESTMENT_AMOUNT, help="Investment amount used for compound Excel/report calculations.")
    parser.add_argument("--timeframes", type=str, default="2h,4h,1day", help="Comma-separated timeframes, e.g. 2h or 4h,1day")
    parser.add_argument("--trials", type=int, default=300)
    parser.add_argument("--jobs", type=int, default=1)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output", type=str, default="results")
    parser.add_argument("--min-trades", type=int, default=12)
    parser.add_argument("--adjust", type=str, default="splits", choices=["all", "splits", "dividends", "none"])
    parser.add_argument("--exchange", type=str, default=None, help="Optional exchange filter, e.g. NASDAQ")
    parser.add_argument("--api-key", type=str, default=None, help="Optional Twelve Data API key. If empty, the built-in DEFAULT_TWELVEDATA_API_KEY is used.")
    parser.add_argument("--pause-sec", type=float, default=0.35, help="Seconds to wait between Twelve Data requests. Increase this for free API limits.")
    parser.add_argument("--quiet-params", action="store_true", help="Do not print all best parameters to console")
    parser.add_argument("--no-json", action="store_true", help="Do not save JSON output")
    args, _unknown = parser.parse_known_args()

    api_key = args.api_key or os.environ.get("TWELVEDATA_API_KEY") or DEFAULT_TWELVEDATA_API_KEY
    if not api_key:
        raise SystemExit("Missing API key. Use --api-key or set TWELVEDATA_API_KEY.")
    if args.years <= 0:
        raise SystemExit("--years must be positive.")
    if args.investment_amount <= 0:
        raise SystemExit("--investment-amount must be positive.")
    try:
        resolve_date_range(args.years, args.start_date, args.end_date)
    except Exception as exc:
        raise SystemExit(f"Invalid analysis date range: {exc}")

    mode = args.mode
    if mode is None:
        mode = "replay" if (args.replay or args.params_file) else "optimize"

    if mode == "replay":
        run_replay_mode(args)
        return

    if mode == "signal":
        run_signal_mode(args)
        return

    if mode == "scan":
        run_scan_mode(args)
        return

    # Optimize mode
    if not args.ticker:
        raise SystemExit("Optimize mode requires a ticker symbol.")

    try:
        selected_timeframes = parse_timeframes(args.timeframes)
    except ValueError as exc:
        raise SystemExit(str(exc))

    os.makedirs(args.output, exist_ok=True)

    cfg = DownloadConfig(
        api_key=api_key,
        symbol=args.ticker.upper(),
        years=args.years,
        start_date=args.start_date,
        end_date=args.end_date,
        adjust=args.adjust,
        exchange=args.exchange,
        pause_sec=max(float(getattr(args, "pause_sec", 0.35) or 0.35), 0.0),
    )
    client = TwelveDataClient(cfg)

    print(f"\nOptimize mode for {cfg.symbol}")
    print(f"Timeframes: {selected_timeframes}")
    print(f"Analysis period: {analysis_period_label(args.years, args.start_date, args.end_date)}")
    print(f"Investment amount: ${args.investment_amount:,.2f}")
    print(f"\nLoading data for {cfg.symbol}...")
    datasets = prepare_datasets(client, selected_timeframes)

    study = run_optimization(
        datasets,
        min_trades=args.min_trades,
        n_trials=args.trials,
        seed=args.seed,
        n_jobs=args.jobs,
        selected_timeframes=selected_timeframes,
    )

    summary, trades, params = summarize_best(study, datasets, cfg.symbol, args.years, args.adjust, investment_amount=args.investment_amount, start_date=args.start_date, end_date=args.end_date)
    print_report(summary, params, show_params=not args.quiet_params)

    top_trials_df = build_top_trials(study)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(args.output, f"{cfg.symbol}_best_strategy_{timestamp}.xlsx")
    json_path = os.path.join(args.output, f"{cfg.symbol}_best_strategy_{timestamp}.json")

    export_excel(excel_path, summary, params, trades, top_trials_df)

    if not args.no_json:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"summary": summary, "params": params}, f, indent=2, default=str)

    print("\nSaved files:")
    print(f"  Excel : {excel_path}")
    if not args.no_json:
        print(f"  JSON  : {json_path}")

    if summary["has_open_trade"] and not trades.empty:
        last_trade = trades.iloc[-1]
        print(
            f"\nOpen trade note: entered at {last_trade['entry_price']:.4f} on {last_trade['entry_time']} "
            f"and is still OPEN. Current marked PnL = {last_trade['pnl']:+.2f} ({last_trade['pnl_pct']:+.2f}%)."
        )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("Interrupted by user.")

